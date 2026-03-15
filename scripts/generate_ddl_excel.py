#!/usr/bin/env python
"""
Generate DDL Excel file for vowsls database schema.
Extracts table and column information from ORM models and creates a comprehensive Excel file.
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple
import importlib.util
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Table schema metadata extracted from models
TABLES_SCHEMA = {
    # ===== MASTER TABLES (mst) =====
    "approval_mst": {
        "description": "Approval configuration master - defines approval levels and limits per menu/branch/user",
        "columns": [
            ("approval_mst_id", "int", "auto_increment", "Primary Key"),
            ("menu_id", "int", "FK -> menu_mst", "Foreign Key"),
            ("user_id", "int", "FK -> user_mst", "Foreign Key"),
            ("branch_id", "int", "FK -> branch_mst", "Foreign Key"),
            ("approval_level", "int", "", "Approval level number"),
            ("max_amount_single", "double", "", "Max approval amount for single transaction"),
            ("day_max_amount", "double", "", "Max approval amount per day"),
            ("month_max_amount", "double", "", "Max approval amount per month"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("updated_by", "int", "", "User ID who updated"),
        ]
    },
    "branch_mst": {
        "description": "Branch master - company branch locations with address and GST details",
        "columns": [
            ("branch_id", "int", "auto_increment", "Primary Key"),
            ("branch_name", "varchar", "255", "Branch name (unique)"),
            ("co_id", "int", "FK -> co_mst", "Company ID"),
            ("branch_address1", "varchar", "255", "Primary address"),
            ("branch_address2", "varchar", "255", "Secondary address"),
            ("branch_zipcode", "int", "", "Postal code"),
            ("country_id", "int", "FK -> country_mst", "Country ID"),
            ("state_id", "int", "FK -> state_mst", "State ID"),
            ("gst_no", "varchar", "25", "GST registration number"),
            ("contact_no", "int", "", "Contact number"),
            ("contact_person", "varchar", "255", "Contact person name"),
            ("branch_email", "varchar", "255", "Email address"),
            ("active", "boolean", "", "Active status"),
            ("gst_verified", "boolean", "", "GST verification flag"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("updated_by", "int", "", "User ID who updated"),
            ("branch_prefix", "varchar", "100", "Prefix for document numbering"),
        ]
    },
    "co_mst": {
        "description": "Company master - main company/tenant configuration",
        "columns": [
            ("co_id", "int", "auto_increment", "Primary Key"),
            ("co_name", "varchar", "255", "Company name (unique)"),
            ("co_prefix", "varchar", "25", "Company prefix (unique)"),
            ("co_address1", "varchar", "255", "Primary address"),
            ("co_address2", "varchar", "255", "Secondary address"),
            ("co_zipcode", "int", "", "Postal code"),
            ("country_id", "int", "FK -> country_mst", "Country ID"),
            ("state_id", "int", "FK -> state_mst", "State ID"),
            ("city_id", "int", "FK -> city_mst", "City ID"),
            ("co_logo", "varchar", "255", "Logo file path"),
            ("auto_datetime_insert", "datetime", "", "Creation timestamp"),
            ("created_by_con_user", "int", "", "Console user who created"),
            ("co_cin_no", "varchar", "25", "CIN registration number"),
            ("co_email_id", "varchar", "255", "Company email"),
            ("co_pan_no", "varchar", "25", "PAN number"),
            ("s3bucket_name", "varchar", "255", "AWS S3 bucket"),
            ("s3folder_name", "varchar", "255", "AWS S3 folder"),
            ("tally_sync", "varchar", "255", "Tally sync flag"),
            ("alert_email_id", "varchar", "255", "Alert email"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "cost_factor_mst": {
        "description": "Cost factor master - cost allocation factors for inventory/production",
        "columns": [
            ("cost_factor_id", "int", "auto_increment", "Primary Key"),
            ("cost_factor_name", "varchar", "255", "Cost factor name"),
            ("cost_factor_desc", "varchar", "255", "Cost factor description"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("dept_id", "int", "FK -> dept_mst", "Department ID (optional)"),
        ]
    },
    "country_mst": {
        "description": "Country master - list of countries",
        "columns": [
            ("country_id", "int", "auto_increment", "Primary Key"),
            ("country", "varchar", "255", "Country name (unique)"),
        ]
    },
    "currency_mst": {
        "description": "Currency master - currency definitions",
        "columns": [
            ("currency_id", "int", "auto_increment", "Primary Key"),
            ("currency_prefix", "varchar", "25", "Currency symbol/prefix"),
        ]
    },
    "dept_mst": {
        "description": "Department master - company departments by branch",
        "columns": [
            ("dept_id", "int", "auto_increment", "Primary Key"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("created_by", "int", "", "User ID who created"),
            ("dept_desc", "varchar", "30", "Department description"),
            ("dept_code", "varchar", "30", "Department code"),
            ("order_id", "int", "", "Display order"),
            ("created_date", "datetime", "", "Creation timestamp"),
        ]
    },
    "entity_type_mst": {
        "description": "Entity type master - business entity classifications",
        "columns": [
            ("entity_type_id", "int", "PK", "Primary Key"),
            ("entity_type_name", "varchar", "30", "Entity type name"),
        ]
    },
    "expense_type_mst": {
        "description": "Expense type master - types of expenses for cost tracking",
        "columns": [
            ("expense_type_id", "int", "auto_increment", "Primary Key"),
            ("expense_type_name", "varchar", "45", "Expense type name"),
            ("active", "int", "", "Active status (1=active, 0=inactive)"),
        ]
    },
    "item_grp_mst": {
        "description": "Item group master - hierarchical item categorization",
        "columns": [
            ("item_grp_id", "bigint", "auto_increment", "Primary Key"),
            ("parent_grp_id", "bigint", "FK -> item_grp_mst", "Parent group ID (hierarchy)"),
            ("active", "varchar", "255", "Active status"),
            ("co_id", "int", "FK -> co_mst", "Company ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("item_grp_name", "varchar", "255", "Group name"),
            ("item_grp_code", "varchar", "255", "Group code"),
            ("purchase_code", "bigint", "", "Purchase classification code"),
            ("item_type_id", "int", "FK -> item_type_master", "Item type ID"),
        ]
    },
    "item_minmax_mst": {
        "description": "Item min/max master - inventory reorder levels per item/branch",
        "columns": [
            ("item_minmax_id", "int", "auto_increment", "Primary Key"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("minqty", "double", "", "Minimum stock quantity"),
            ("maxqty", "double", "", "Maximum stock quantity"),
            ("min_order_qty", "double", "", "Minimum order quantity"),
            ("lead_time", "int", "", "Lead time in days"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "item_mst": {
        "description": "Item master table - individual items/products/materials",
        "columns": [
            ("item_id", "int", "auto_increment", "Primary Key"),
            ("active", "int", "", "Active status (1=active, 0=inactive)"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("updated_by", "int", "", "User ID who updated"),
            ("item_grp_id", "bigint", "FK -> item_grp_mst", "Item group ID"),
            ("item_code", "varchar", "255", "Item code"),
            ("tangible", "boolean", "", "Tangible item flag"),
            ("item_name", "varchar", "255", "Item name"),
            ("item_photo", "text", "", "Item photo/image data"),
            ("legacy_item_code", "varchar", "255", "Legacy system item code"),
            ("hsn_code", "varchar", "255", "HSN/SAC code"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("tax_percentage", "double", "", "Default tax percentage"),
            ("saleable", "boolean", "", "Can be sold flag"),
            ("consumable", "boolean", "", "Consumable item flag"),
            ("purchaseable", "boolean", "", "Can be purchased flag"),
            ("manufacturable", "boolean", "", "Can be manufactured flag"),
            ("assembly", "boolean", "", "Assembly item flag"),
            ("uom_rounding", "int", "", "UOM rounding decimals"),
            ("rate_rounding", "int", "", "Rate rounding decimals"),
        ]
    },
    "item_type_master": {
        "description": "Item type master table",
        "columns": [
            ("item_type_id", "int", "auto_increment", "Primary Key"),
            ("item_type_name", "varchar", "25", "Item type name"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "item_make": {
        "description": "Item make/brand master table - manufacturer/brand info for items",
        "columns": [
            ("item_make_id", "int", "auto_increment", "Primary Key"),
            ("item_make_name", "varchar", "255", "Make/brand name"),
            ("item_grp_id", "bigint", "FK -> item_grp_mst", "Item group ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "menu_mst": {
        "description": "Menu master - application menu/navigation item definitions",
        "columns": [
            ("menu_id", "int", "auto_increment", "Primary Key"),
            ("menu_name", "varchar", "255", "Menu item name (unique)"),
            ("menu_path", "varchar", "255", "URL/route path"),
            ("active", "boolean", "", "Active status"),
            ("menu_parent_id", "int", "FK -> menu_mst", "Parent menu ID (hierarchy)"),
            ("menu_type_id", "int", "FK -> menu_type_mst", "Menu type ID"),
            ("menu_icon", "varchar", "255", "Icon name/path"),
        ]
    },
    "menu_type_mst": {
        "description": "Menu type master - types of menu items",
        "columns": [
            ("menu_type_id", "int", "auto_increment", "Primary Key"),
            ("menu_type", "varchar", "25", "Menu type name (unique)"),
        ]
    },
    "party_mst": {
        "description": "Party master - customers, suppliers, vendors",
        "columns": [
            ("party_id", "int", "auto_increment", "Primary Key"),
            ("party_name", "varchar", "255", "Party name"),
            ("party_type_id", "int", "FK -> party_type_mst", "Party type ID"),
            ("party_code", "varchar", "50", "Party code"),
            ("entity_type_id", "int", "FK -> entity_type_mst", "Entity type ID"),
            ("active", "int", "", "Active status"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("co_id", "int", "FK -> co_mst", "Company ID"),
        ]
    },
    "party_branch_mst": {
        "description": "Party branch master - branch/location details for parties",
        "columns": [
            ("party_branch_id", "int", "auto_increment", "Primary Key"),
            ("party_id", "int", "FK -> party_mst", "Party ID"),
            ("branch_name", "varchar", "255", "Branch name"),
            ("address1", "varchar", "255", "Primary address"),
            ("address2", "varchar", "255", "Secondary address"),
            ("city", "varchar", "100", "City"),
            ("state", "varchar", "100", "State"),
            ("country", "varchar", "100", "Country"),
            ("zipcode", "varchar", "20", "Postal code"),
            ("contact_person", "varchar", "255", "Contact person name"),
            ("contact_no", "varchar", "20", "Contact number"),
            ("email", "varchar", "255", "Email address"),
            ("gst_no", "varchar", "25", "GST number"),
            ("bank_details", "text", "", "Bank details"),
            ("factory_address", "text", "", "Factory address"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "party_type_mst": {
        "description": "Party type master - categorizes parties (supplier, customer, etc.)",
        "columns": [
            ("party_type_id", "int", "auto_increment", "Primary Key"),
            ("party_type_name", "varchar", "50", "Party type name"),
            ("active", "int", "", "Active status"),
        ]
    },
    "roles_mst": {
        "description": "Roles master - user roles and permissions",
        "columns": [
            ("role_id", "int", "auto_increment", "Primary Key"),
            ("role_name", "varchar", "255", "Role name (unique)"),
            ("active", "boolean", "", "Active status"),
            ("created_by_con_user", "int", "", "Console user who created"),
            ("created_date_time", "datetime", "", "Creation timestamp"),
        ]
    },
    "state_mst": {
        "description": "State master - list of states/provinces",
        "columns": [
            ("state_id", "int", "auto_increment", "Primary Key"),
            ("state", "varchar", "255", "State name (unique)"),
            ("country_id", "int", "FK -> country_mst", "Country ID"),
        ]
    },
    "uom_mst": {
        "description": "Unit of measure master - measurement units (kg, litre, pieces, etc.)",
        "columns": [
            ("uom_id", "int", "auto_increment", "Primary Key"),
            ("uom_name", "varchar", "25", "Unit of measure name (unique)"),
            ("uom_desc", "varchar", "100", "Description"),
        ]
    },
    "user_mst": {
        "description": "User master - portal users",
        "columns": [
            ("user_id", "int", "auto_increment", "Primary Key"),
            ("email_id", "varchar", "255", "Email address (unique)"),
            ("name", "varchar", "255", "User full name"),
            ("password", "varchar", "255", "Password hash"),
            ("refresh_token", "varchar", "255", "JWT refresh token"),
            ("active", "boolean", "", "Active status"),
            ("created_by_con_user", "int", "", "Console user who created"),
            ("created_date_time", "datetime", "", "Creation timestamp"),
        ]
    },
    "user_role_map": {
        "description": "User-role mapping - assigns roles to users per company/branch",
        "columns": [
            ("user_role_map_id", "bigint", "auto_increment", "Primary Key"),
            ("user_id", "int", "FK -> user_mst", "User ID"),
            ("role_id", "int", "FK -> roles_mst", "Role ID"),
            ("co_id", "int", "FK -> co_mst", "Company ID"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("created_by_con_user", "int", "", "Console user who created"),
            ("created_at", "datetime", "", "Creation timestamp"),
        ]
    },
    "role_menu_map": {
        "description": "Role-menu mapping - assigns menu items to roles with access types",
        "columns": [
            ("role_menu_mapping_id", "int", "auto_increment", "Primary Key"),
            ("role_id", "int", "FK -> roles_mst", "Role ID"),
            ("menu_id", "int", "FK -> menu_mst", "Menu ID"),
            ("access_type_id", "int", "FK -> access_type", "Access type ID"),
        ]
    },
    "access_type": {
        "description": "Access type master - defines permission levels",
        "columns": [
            ("access_type_id", "int", "auto_increment", "Primary Key"),
            ("access_type", "varchar", "25", "Access type (unique)"),
        ]
    },

    # ===== PROCUREMENT TABLES =====
    "proc_indent": {
        "description": "Purchase indent header table",
        "columns": [
            ("indent_id", "int", "auto_increment", "Primary Key"),
            ("indent_date", "date", "", "Indent creation date"),
            ("indent_no", "int", "", "Indent number"),
            ("active", "boolean", "", "Active status"),
            ("indent_type_id", "varchar", "25", "Indent type (Regular/BOM/Open)"),
            ("remarks", "varchar", "500", "Remarks"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("expense_type_id", "int", "FK -> expense_type_mst", "Expense type ID"),
            ("project_id", "int", "", "Project ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("status_id", "int", "", "Status ID (21=Draft, 1=Open, 20=Pending, 3=Approved, etc.)"),
            ("indent_title", "varchar", "255", "Indent title"),
            ("dept_id", "int", "FK -> dept_mst", "Department ID"),
            ("approval_level", "int", "", "Current approval level"),
        ]
    },
    "proc_indent_dtl": {
        "description": "Purchase indent detail/line items table",
        "columns": [
            ("indent_dtl_id", "int", "auto_increment", "Primary Key"),
            ("indent_id", "int", "FK -> proc_indent", "Indent ID"),
            ("required_by_days", "int", "", "Days until required"),
            ("active", "boolean", "", "Active status"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("qty", "double", "", "Quantity required"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("remarks", "varchar", "599", "Line item remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("item_make_id", "int", "FK -> item_make", "Item make/brand ID"),
            ("dept_id", "int", "FK -> dept_mst", "Department ID"),
            ("state", "int", "", "State flag"),
        ]
    },
    "proc_indent_dtl_cancel": {
        "description": "Cancelled indent detail line items",
        "columns": [
            ("indent_dtl_cancel_id", "int", "auto_increment", "Primary Key"),
            ("indent_dtl_id", "int", "FK -> proc_indent_dtl", "Indent detail ID"),
            ("cancelled_by", "int", "", "User ID who cancelled"),
            ("cancelled_date_time", "datetime", "", "Cancellation timestamp"),
            ("cancelled_qty", "double", "", "Quantity cancelled"),
            ("cancelled_reasons", "varchar", "500", "Cancellation reason"),
        ]
    },
    "proc_po": {
        "description": "Purchase order header table",
        "columns": [
            ("po_id", "int", "auto_increment", "Primary Key"),
            ("po_date", "date", "", "Purchase order date"),
            ("po_no", "int", "", "Purchase order number"),
            ("active", "boolean", "", "Active status"),
            ("supplier_id", "int", "FK -> party_mst", "Supplier/party ID"),
            ("supplier_branch_id", "int", "FK -> party_branch_mst", "Supplier branch ID"),
            ("branch_id", "int", "FK -> branch_mst", "Company branch ID"),
            ("delivery_date", "date", "", "Expected delivery date"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("status_id", "int", "", "Status ID"),
            ("payment_terms", "varchar", "255", "Payment terms"),
            ("dept_id", "int", "FK -> dept_mst", "Department ID"),
            ("approval_level", "int", "", "Current approval level"),
        ]
    },
    "proc_po_dtl": {
        "description": "Purchase order detail/line items",
        "columns": [
            ("po_dtl_id", "int", "auto_increment", "Primary Key"),
            ("po_id", "int", "FK -> proc_po", "Purchase order ID"),
            ("indent_dtl_id", "int", "FK -> proc_indent_dtl", "Linked indent detail ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("qty", "double", "", "Order quantity"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("rate", "double", "", "Unit rate"),
            ("amount", "double", "", "Line total amount"),
            ("remarks", "varchar", "500", "Line remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("item_make_id", "int", "FK -> item_make", "Item make ID"),
        ]
    },
    "proc_po_gst": {
        "description": "GST details for PO line items",
        "columns": [
            ("po_gst_id", "int", "auto_increment", "Primary Key"),
            ("po_dtl_id", "int", "FK -> proc_po_dtl", "PO detail line ID"),
            ("tax_pct", "double", "", "GST tax percentage"),
            ("cgst_amount", "double", "", "CGST amount"),
            ("sgst_amount", "double", "", "SGST amount"),
            ("igst_amount", "double", "", "IGST amount"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "proc_po_additional": {
        "description": "Additional charges for purchase order",
        "columns": [
            ("po_additional_id", "int", "auto_increment", "Primary Key"),
            ("po_id", "int", "FK -> proc_po", "Purchase order ID"),
            ("charge_type", "varchar", "100", "Type of charge"),
            ("charge_amount", "double", "", "Charge amount"),
            ("remarks", "varchar", "255", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "proc_inward": {
        "description": "Goods receipt/inward header table",
        "columns": [
            ("inward_id", "int", "auto_increment", "Primary Key"),
            ("inward_date", "date", "", "Goods receipt date"),
            ("inward_no", "int", "", "Inward number"),
            ("po_id", "int", "FK -> proc_po", "Purchase order ID"),
            ("branch_id", "int", "FK -> branch_mst", "Receiving branch ID"),
            ("supplier_id", "int", "FK -> party_mst", "Supplier ID"),
            ("supplier_branch_id", "int", "FK -> party_branch_mst", "Supplier branch ID"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("approval_level", "int", "", "Current approval level"),
        ]
    },
    "proc_inward_dtl": {
        "description": "Goods receipt detail/line items",
        "columns": [
            ("inward_dtl_id", "int", "auto_increment", "Primary Key"),
            ("inward_id", "int", "FK -> proc_inward", "Inward ID"),
            ("po_dtl_id", "int", "FK -> proc_po_dtl", "PO detail ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("received_qty", "double", "", "Quantity received"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("rate", "double", "", "Unit rate"),
            ("amount", "double", "", "Line total amount"),
            ("remarks", "varchar", "500", "Remarks"),
            ("warehouse_id", "int", "FK -> warehouse_mst", "Warehouse ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "proc_gst": {
        "description": "GST details for procurement inward items",
        "columns": [
            ("gst_invoice_type", "int", "auto_increment", "Primary Key"),
            ("proc_inward_dtl", "int", "FK -> proc_inward_dtl", "Inward detail ID"),
            ("tax_pct", "double", "", "GST tax percentage"),
            ("s_tax_amount", "double", "", "SGST amount"),
            ("i_tax_amount", "double", "", "IGST amount"),
            ("c_tax_amount", "double", "", "CGST amount"),
            ("tax_amount", "double", "", "Total tax amount"),
            ("active", "int", "", "Active status"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "proc_enquiry": {
        "description": "Price enquiry header table",
        "columns": [
            ("enquiry_id", "int", "auto_increment", "Primary Key"),
            ("price_enquiry_date", "date", "", "Enquiry date"),
            ("price_enquiry_squence_no", "varchar", "30", "Sequence number"),
            ("remarks", "varchar", "255", "Remarks"),
            ("status_id", "int", "", "Status ID"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("suppliers", "varchar", "255", "Supplier list"),
            ("terms_conditions", "varchar", "255", "Terms & conditions"),
            ("delivery_days", "int", "", "Delivery days"),
            ("active", "int", "", "Active status"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "proc_enquiry_dtl": {
        "description": "Price enquiry detail/line items",
        "columns": [
            ("enquiry_dtl_id", "int", "auto_increment", "Primary Key"),
            ("enquiry_id", "int", "FK -> proc_enquiry", "Enquiry ID"),
            ("indent_dtl_id", "int", "FK -> proc_indent_dtl", "Indent detail ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("item_make_id", "int", "FK -> item_make", "Item make ID"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("qty", "double", "", "Quantity"),
            ("remarks", "varchar", "255", "Remarks"),
            ("active", "int", "", "Active status"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },

    # ===== SALES TABLES =====
    "sales_invoice": {
        "description": "Sales invoice header table",
        "columns": [
            ("invoice_id", "bigint", "auto_increment", "Primary Key"),
            ("invoice_no", "int", "", "Invoice number"),
            ("invoice_date", "date", "", "Invoice date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("customer_id", "int", "FK -> party_mst", "Customer ID"),
            ("customer_branch_id", "int", "FK -> party_branch_mst", "Customer branch ID"),
            ("delivery_order_id", "bigint", "FK -> sales_delivery_order", "Linked delivery order"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("approval_level", "int", "", "Current approval level"),
        ]
    },
    "invoice_line_item": {
        "description": "Sales invoice detail/line items",
        "columns": [
            ("invoice_li_id", "bigint", "auto_increment", "Primary Key"),
            ("invoice_id", "bigint", "FK -> sales_invoice", "Invoice ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("quantity", "double", "", "Quantity invoiced"),
            ("uom", "varchar", "255", "Unit of measure"),
            ("rate", "double", "", "Unit rate"),
            ("amount", "double", "", "Line total amount"),
            ("bales", "varchar", "100", "Bales information"),
            ("is_active", "int", "", "Active status"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_delivery_order": {
        "description": "Sales delivery order header",
        "columns": [
            ("delivery_order_id", "bigint", "auto_increment", "Primary Key"),
            ("do_no", "int", "", "Delivery order number"),
            ("do_date", "date", "", "Delivery date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("customer_id", "int", "FK -> party_mst", "Customer ID"),
            ("customer_branch_id", "int", "FK -> party_branch_mst", "Customer branch ID"),
            ("sales_order_id", "bigint", "FK -> sales_order", "Linked sales order"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_delivery_order_dtl": {
        "description": "Sales delivery order detail/line items",
        "columns": [
            ("do_dtl_id", "bigint", "auto_increment", "Primary Key"),
            ("delivery_order_id", "bigint", "FK -> sales_delivery_order", "Delivery order ID"),
            ("sales_order_dtl_id", "bigint", "FK -> sales_order_dtl", "Sales order detail ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("qty", "double", "", "Delivery quantity"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_order": {
        "description": "Sales order header",
        "columns": [
            ("sales_order_id", "bigint", "auto_increment", "Primary Key"),
            ("so_no", "int", "", "Sales order number"),
            ("so_date", "date", "", "Sales order date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("customer_id", "int", "FK -> party_mst", "Customer ID"),
            ("customer_branch_id", "int", "FK -> party_branch_mst", "Customer branch ID"),
            ("quotation_id", "bigint", "FK -> sales_quotation", "Linked quotation"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_order_dtl": {
        "description": "Sales order detail/line items",
        "columns": [
            ("so_dtl_id", "bigint", "auto_increment", "Primary Key"),
            ("sales_order_id", "bigint", "FK -> sales_order", "Sales order ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("qty", "double", "", "Order quantity"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("rate", "double", "", "Unit rate"),
            ("amount", "double", "", "Line total amount"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_quotation": {
        "description": "Sales quotation header",
        "columns": [
            ("quotation_id", "bigint", "auto_increment", "Primary Key"),
            ("quotation_no", "int", "", "Quotation number"),
            ("quotation_date", "date", "", "Quotation date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("customer_id", "int", "FK -> party_mst", "Customer ID"),
            ("customer_branch_id", "int", "FK -> party_branch_mst", "Customer branch ID"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "sales_quotation_dtl": {
        "description": "Sales quotation detail/line items",
        "columns": [
            ("quotation_lineitem_id", "int", "auto_increment", "Primary Key"),
            ("sales_quotation_id", "bigint", "FK -> sales_quotation", "Quotation ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("quantity", "double", "", "Quote quantity"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("rate", "double", "", "Unit rate"),
            ("amount", "double", "", "Line total amount"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },

    # ===== INVENTORY TABLES =====
    "issue_hdr": {
        "description": "Inventory issue header table",
        "columns": [
            ("issue_id", "int", "auto_increment", "Primary Key"),
            ("issue_no", "int", "", "Issue number"),
            ("issue_date", "date", "", "Issue date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("warehouse_from_id", "int", "FK -> warehouse_mst", "Source warehouse ID"),
            ("warehouse_to_id", "int", "FK -> warehouse_mst", "Destination warehouse ID (if transfer)"),
            ("dept_id", "int", "FK -> dept_mst", "Department ID"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
    "issue_li": {
        "description": "Inventory issue line items",
        "columns": [
            ("issue_li_id", "int", "auto_increment", "Primary Key"),
            ("issue_id", "int", "FK -> issue_hdr", "Issue ID"),
            ("item_id", "int", "FK -> item_mst", "Item ID"),
            ("uom_id", "int", "FK -> uom_mst", "Unit of measure ID"),
            ("req_quantity", "double", "", "Requested quantity"),
            ("issue_qty", "double", "", "Actual issued quantity"),
            ("expense_type_id", "int", "FK -> expense_type_mst", "Expense type ID"),
            ("cost_factor_id", "int", "FK -> cost_factor_mst", "Cost factor ID"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },

    # ===== WAREHOUSE TABLES =====
    "warehouse_mst": {
        "description": "Warehouse master - warehouse locations",
        "columns": [
            ("warehouse_id", "int", "auto_increment", "Primary Key"),
            ("warehouse_name", "varchar", "30", "Warehouse name"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("warehouse_type", "varchar", "20", "Type (Main, Sub, etc.)"),
            ("parent_warehouse_id", "int", "FK -> warehouse_mst", "Parent warehouse (hierarchy)"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
            ("updated_by", "int", "", "User ID who updated"),
        ]
    },

    # ===== JUTE TABLES (Sample - expand as needed) =====
    "jute_mr": {
        "description": "Jute material receipt header",
        "columns": [
            ("jute_mr_id", "bigint", "auto_increment", "Primary Key"),
            ("mr_no", "int", "", "Material receipt number"),
            ("mr_date", "date", "", "MR date"),
            ("branch_id", "int", "FK -> branch_mst", "Branch ID"),
            ("supplier_id", "int", "FK -> party_mst", "Supplier ID"),
            ("po_id", "int", "FK -> proc_po", "Purchase order ID"),
            ("status_id", "int", "", "Status ID"),
            ("remarks", "varchar", "500", "Remarks"),
            ("updated_by", "int", "", "User ID who updated"),
            ("updated_date_time", "datetime", "", "Last updated timestamp"),
        ]
    },
}

def create_ddl_excel(output_path: str = "vowsls_ddl_schema.xlsx"):
    """
    Create Excel file with DDL information for all tables in vowsls database.
    
    Args:
        output_path: Path to output Excel file
    """
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "DDL Schema"
    
    # Define headers
    headers = ["Table Name", "Table Description", "Column Name", "Data Type", "Size/Constraint", "Description"]
    
    # Set header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    
    # Prepare data
    rows = []
    row_num = 2
    
    for table_name, table_info in TABLES_SCHEMA.items():
        description = table_info.get("description", "")
        columns = table_info.get("columns", [])
        
        for idx, col in enumerate(columns):
            if len(col) == 4:
                col_name, data_type, size_or_constraint, col_description = col
            else:
                col_name, data_type, size_or_constraint, col_description = col[0], col[1], col[2] if len(col) > 2 else "", col[3] if len(col) > 3 else ""
            
            # Write row data
            ws.cell(row=row_num, column=1).value = table_name if idx == 0 else ""
            ws.cell(row=row_num, column=2).value = description if idx == 0 else ""
            ws.cell(row=row_num, column=3).value = col_name
            ws.cell(row=row_num, column=4).value = data_type
            ws.cell(row=row_num, column=5).value = size_or_constraint
            ws.cell(row=row_num, column=6).value = col_description
            
            # Apply alternating row colors
            if row_num % 2 == 0:
                light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                for col in range(1, 7):
                    ws.cell(row=row_num, column=col).fill = light_fill
            
            # Set text wrapping and alignment
            for col in [1, 2, 3, 4, 5, 6]:
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal="left" if col in [2, 6] else "center", vertical="top", wrap_text=True)
            
            row_num += 1
    
    # Add a summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws.column_dimensions['A'].width = 40
    summary_ws.column_dimensions['B'].width = 15
    
    # Summary headers
    summary_ws['A1'] = "Database DDL Schema Summary"
    summary_ws['A1'].font = Font(bold=True, size=14)
    
    # Summary info
    summary_ws['A3'] = "Database Name"
    summary_ws['B3'] = "vowsls"
    summary_ws['A4'] = "Total Tables"
    summary_ws['B4'] = len(TABLES_SCHEMA)
    summary_ws['A5'] = "Table Categories"
    summary_ws['B5'] = "Master, Procurement, Sales, Inventory, Jute"
    summary_ws['A6'] = "Generated Date"
    from datetime import datetime
    summary_ws['B6'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Table count by category
    summary_ws['A8'] = "Table Breakdown by Category"
    summary_ws['A8'].font = Font(bold=True, size=11)
    
    summary_ws['A9'] = "Category"
    summary_ws['B9'] = "Count"
    summary_ws['A9'].font = Font(bold=True)
    summary_ws['B9'].font = Font(bold=True)
    
    categories = {
        "Master Tables": 0,
        "Procurement Tables": 0,
        "Sales Tables": 0,
        "Inventory Tables": 0,
        "Warehouse Tables": 0,
        "Jute Tables": 0,
    }
    
    master_keywords = ["_mst", "role_menu_map", "access_type", "user_role_map", "menu_mst", "approval_mst"]
    proc_keywords = ["proc_"]
    sales_keywords = ["sales_", "invoice_", "sales_quotation", "sales_delivery_order"]
    inventory_keywords = ["issue_", "warehouse_mst"]
    jute_keywords = ["jute_"]
    
    for table in TABLES_SCHEMA.keys():
        if any(k in table for k in master_keywords):
            categories["Master Tables"] += 1
        elif any(k in table for k in proc_keywords):
            categories["Procurement Tables"] += 1
        elif any(k in table for k in sales_keywords):
            categories["Sales Tables"] += 1
        elif any(k in table for k in jute_keywords):
            categories["Jute Tables"] += 1
        elif "warehouse" in table:
            categories["Warehouse Tables"] += 1
        else:
            categories["Inventory Tables"] += 1
    
    row_idx = 10
    for category, count in categories.items():
        if count > 0:
            summary_ws[f'A{row_idx}'] = category
            summary_ws[f'B{row_idx}'] = count
            row_idx += 1
    
    # Save workbook
    wb.save(output_path)
    
    print(f"✓ DDL Excel file created: {output_path}")
    print(f"✓ Total tables: {len(TABLES_SCHEMA)}")
    total_rows = sum(len(t.get("columns", [])) for t in TABLES_SCHEMA.values())
    print(f"✓ Total columns: {total_rows}")

if __name__ == "__main__":
    output_file = "vowsls_ddl_schema.xlsx"
    create_ddl_excel(output_file)
    print(f"\nFile location: {Path(output_file).absolute()}")
