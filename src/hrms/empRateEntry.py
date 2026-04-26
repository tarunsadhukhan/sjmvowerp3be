"""HRMS Employee Rate Entry endpoints."""

import base64
import io
from datetime import datetime, date

from fastapi import APIRouter, Depends, File, HTTPException, Request, Response, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy.sql import text

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import get_tenant_db

router = APIRouter()


@router.get("/emp_rate_employee_lookup")
async def emp_rate_employee_lookup(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Validate emp_code and return eb_id, employee name and branch_id."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        emp_code = request.query_params.get("emp_code")
        if not emp_code:
            raise HTTPException(status_code=400, detail="emp_code is required")

        lookup_query = text(
            """
            SELECT
                o.eb_id,
                o.emp_code,
                o.branch_id,
                CONCAT(
                    COALESCE(p.first_name, ''),
                    CASE WHEN p.middle_name IS NULL OR p.middle_name = '' THEN '' ELSE CONCAT(' ', p.middle_name) END,
                    CASE WHEN p.last_name IS NULL OR p.last_name = '' THEN '' ELSE CONCAT(' ', p.last_name) END
                ) AS employee_name
            FROM hrms_ed_official_details o
            INNER JOIN hrms_ed_personal_details p ON p.eb_id = o.eb_id
            WHERE o.emp_code = :emp_code
              AND o.active = 1
              AND p.active = 1
            LIMIT 1
            """
        )

        row = db.execute(lookup_query, {"emp_code": emp_code.strip()}).fetchone()
        if not row:
            return {"found": False, "data": None}

        return {"found": True, "data": dict(row._mapping)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/emp_rate_create")
async def emp_rate_create(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Create employee rate entry after validating eb_id exists in official and personal tables."""
    try:
        body = await request.json()

        co_id = body.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        eb_id = body.get("eb_id")
        if not eb_id:
            raise HTTPException(status_code=400, detail="eb_id is required")

        rate = body.get("rate")
        if rate in (None, ""):
            raise HTTPException(status_code=400, detail="rate is required")

        date_of_rate_update = body.get("date_of_rate_update")
        if not date_of_rate_update:
            raise HTTPException(status_code=400, detail="date_of_rate_update is required")

        validate_query = text(
            """
            SELECT o.eb_id
            FROM hrms_ed_official_details o
            INNER JOIN hrms_ed_personal_details p ON p.eb_id = o.eb_id
            WHERE o.eb_id = :eb_id
              AND o.active = 1
              AND p.active = 1
            LIMIT 1
            """
        )
        exists_row = db.execute(validate_query, {"eb_id": int(eb_id)}).fetchone()
        if not exists_row:
            raise HTTPException(
                status_code=400,
                detail="Invalid eb_id: not found in hrms_ed_offical_details/hrms_ed_personal_details",
            )

        try:
            parsed_date = datetime.strptime(str(date_of_rate_update), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date_of_rate_update must be YYYY-MM-DD")

        user_id = token_data.get("user_id") if token_data else None

        insert_query = text(
            """
            INSERT INTO employee_rate_table (
                eb_id,
                rate,
                rate_date,
                updated_date_time
            ) VALUES (
                :eb_id,
                :rate,
                :rate_date,
                NOW()
            )
            """
        )
        db.execute(
            insert_query,
            {
                "eb_id": int(eb_id),
                "rate": int(float(rate)),
                "rate_date": parsed_date,
            },
        )
        _ = user_id  # unused; column not present in employee_rate_table
        db.commit()

        return {"message": "Employee rate saved successfully"}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/emp_rate_list")
async def emp_rate_list(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return paginated list of employee rate entries."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))
        offset = (page - 1) * limit
        search = request.query_params.get("search")

        where_parts: list = []
        params: dict = {"limit": limit, "offset": offset}

        if search:
            where_parts.append(
                "(o.emp_code LIKE :search OR CONCAT(COALESCE(p.first_name,''), ' ', COALESCE(p.last_name,'')) LIKE :search)"
            )
            params["search"] = f"%{search}%"

        where_clause = ("WHERE " + " AND ".join(where_parts)) if where_parts else ""

        base_from = """
            FROM employee_rate_table ert
            INNER JOIN hrms_ed_official_details o ON o.eb_id = ert.eb_id AND o.active = 1
            INNER JOIN hrms_ed_personal_details p ON p.eb_id = ert.eb_id AND p.active = 1
            LEFT JOIN branch_mst b ON b.branch_id = o.branch_id
        """

        list_query = text(
            f"""
            SELECT
                ert.emp_rate_id AS rate_id,
                ert.eb_id,
                o.emp_code,
                CONCAT(
                    COALESCE(p.first_name, ''),
                    CASE WHEN p.middle_name IS NULL OR p.middle_name = '' THEN '' ELSE CONCAT(' ', p.middle_name) END,
                    CASE WHEN p.last_name IS NULL OR p.last_name = '' THEN '' ELSE CONCAT(' ', p.last_name) END
                ) AS employee_name,
                COALESCE(b.branch_name, '') AS branch_name,
                o.branch_id,
                ert.rate,
                DATE_FORMAT(ert.rate_date, '%Y-%m-%d') AS date_of_rate_update
            {base_from}
            {where_clause}
            ORDER BY ert.emp_rate_id DESC
            LIMIT :limit OFFSET :offset
            """
        )

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        count_query = text(f"SELECT COUNT(*) AS total {base_from} {where_clause}")
        total = db.execute(count_query, count_params).scalar() or 0

        rows = db.execute(list_query, params).fetchall()
        data = [dict(r._mapping) for r in rows]

        return {"data": data, "total": total, "page": page, "page_size": limit}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/emp_rate_by_id/{rate_id}")
async def emp_rate_by_id(
    rate_id: int,
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return a single employee rate entry by rate_id."""
    try:
        query = text(
            """
            SELECT
                ert.emp_rate_id AS rate_id,
                ert.eb_id,
                o.emp_code,
                CONCAT(
                    COALESCE(p.first_name, ''),
                    CASE WHEN p.middle_name IS NULL OR p.middle_name = '' THEN '' ELSE CONCAT(' ', p.middle_name) END,
                    CASE WHEN p.last_name IS NULL OR p.last_name = '' THEN '' ELSE CONCAT(' ', p.last_name) END
                ) AS employee_name,
                o.branch_id,
                ert.rate,
                DATE_FORMAT(ert.rate_date, '%Y-%m-%d') AS date_of_rate_update
            FROM employee_rate_table ert
            INNER JOIN hrms_ed_official_details o ON o.eb_id = ert.eb_id AND o.active = 1
            INNER JOIN hrms_ed_personal_details p ON p.eb_id = ert.eb_id AND p.active = 1
            WHERE ert.emp_rate_id = :rate_id
            LIMIT 1
            """
        )
        row = db.execute(query, {"rate_id": rate_id}).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Rate entry not found")
        return {"data": dict(row._mapping)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/emp_rate_update/{rate_id}")
async def emp_rate_update(
    rate_id: int,
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Update rate and date for an existing employee rate entry."""
    try:
        body = await request.json()

        rate = body.get("rate")
        if rate in (None, ""):
            raise HTTPException(status_code=400, detail="rate is required")

        date_of_rate_update = body.get("date_of_rate_update")
        if not date_of_rate_update:
            raise HTTPException(status_code=400, detail="date_of_rate_update is required")

        try:
            parsed_date = datetime.strptime(str(date_of_rate_update), "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(status_code=400, detail="date_of_rate_update must be YYYY-MM-DD")

        user_id = token_data.get("user_id") if token_data else None

        update_query = text(
            """
            UPDATE employee_rate_table
            SET rate = :rate,
                rate_date = :rate_date,
                updated_date_time = NOW()
            WHERE emp_rate_id = :rate_id
            """
        )
        db.execute(
            update_query,
            {
                "rate": int(float(rate)),
                "rate_date": parsed_date,
                "rate_id": rate_id,
            },
        )
        _ = user_id  # unused; column not present in employee_rate_table
        db.commit()

        return {"message": "Employee rate updated successfully"}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Excel bulk upload
# ---------------------------------------------------------------------------

EXPECTED_COLUMNS = ["emp_code", "rate", "rate_date"]


def _parse_rate_date(value):
    """Parse a cell value into a date object. Returns None if invalid."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_rate(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _read_excel_rows(file_bytes: bytes):
    """Return list of dicts with original row_index (1-based, excluding header)
    and raw values for emp_code, rate, rate_date."""
    try:
        import openpyxl  # local import keeps cold start light
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=False)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    header = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    missing = [c for c in EXPECTED_COLUMNS if c not in header]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )
    col_idx = {name: header.index(name) for name in EXPECTED_COLUMNS}

    rows = []
    for r_idx, row_cells in enumerate(rows_iter, start=2):  # data rows start at 2
        values = [c.value for c in row_cells]
        # skip fully blank rows
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in values):
            continue
        emp_code_raw = values[col_idx["emp_code"]] if col_idx["emp_code"] < len(values) else None
        rate_raw = values[col_idx["rate"]] if col_idx["rate"] < len(values) else None
        date_raw = values[col_idx["rate_date"]] if col_idx["rate_date"] < len(values) else None
        rows.append({
            "row_index": r_idx,
            "emp_code": str(emp_code_raw).strip() if emp_code_raw is not None else "",
            "rate_raw": rate_raw,
            "date_raw": date_raw,
        })
    return rows, wb, ws, col_idx


def _build_invalid_workbook(wb, ws, invalid_row_indices, error_messages):
    """Return base64-encoded xlsx with invalid rows highlighted red and an
    appended 'error' column."""
    try:
        from openpyxl.styles import PatternFill
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=str(e))

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")

    # Append "error" header in next available column
    err_col = ws.max_column + 1
    ws.cell(row=1, column=err_col, value="error")

    invalid_set = set(invalid_row_indices)
    for r_idx, msg in error_messages.items():
        ws.cell(row=r_idx, column=err_col, value=msg)
        for c_idx in range(1, err_col + 1):
            ws.cell(row=r_idx, column=c_idx).fill = red_fill
    _ = invalid_set  # kept for symmetry

    out = io.BytesIO()
    wb.save(out)
    return base64.b64encode(out.getvalue()).decode("ascii")


@router.post("/emp_rate_excel_validate")
async def emp_rate_excel_validate(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Parse the uploaded Excel and validate each row.

    Returns:
        {
          new_rows:       [{row_index, emp_code, eb_id, rate, rate_date}],
          existing_rows:  [{row_index, emp_code, eb_id, rate, rate_date,
                             existing_rate_id, existing_rate}],
          invalid_count:  int,
          invalid_file:   base64 xlsx (only when invalid_count > 0),
          invalid_filename: str,
        }
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    rows, wb, ws, _col_idx = _read_excel_rows(file_bytes)
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in Excel")

    # Resolve emp_codes -> eb_id in one query
    emp_codes = list({r["emp_code"] for r in rows if r["emp_code"]})
    emp_map: dict[str, int] = {}
    if emp_codes:
        params = {f"e{i}": code for i, code in enumerate(emp_codes)}
        placeholders = ",".join(f":e{i}" for i in range(len(emp_codes)))
        emp_rows = db.execute(
            text(f"""
                SELECT emp_code, eb_id
                FROM hrms_ed_official_details
                WHERE active = 1 AND emp_code IN ({placeholders})
            """),
            params,
        ).fetchall()
        for er in emp_rows:
            emp_map[str(er.emp_code)] = int(er.eb_id)

    new_rows: list[dict] = []
    existing_rows: list[dict] = []
    error_messages: dict[int, str] = {}

    for r in rows:
        errs: list[str] = []
        emp_code = r["emp_code"]
        rate = _parse_rate(r["rate_raw"])
        rate_date = _parse_rate_date(r["date_raw"])

        if not emp_code:
            errs.append("emp_code is required")
        if rate is None:
            errs.append("rate is invalid")
        if rate_date is None:
            errs.append("rate_date is invalid")

        eb_id = emp_map.get(emp_code) if emp_code else None
        if emp_code and eb_id is None:
            errs.append("emp_code not found in employee master")

        if errs:
            error_messages[r["row_index"]] = "; ".join(errs)
            continue

        # check existing rate row for this eb_id + rate_date
        existing = db.execute(
            text("""
                SELECT emp_rate_id, rate
                FROM employee_rate_table
                WHERE eb_id = :eb_id AND rate_date = :rate_date
                LIMIT 1
            """),
            {"eb_id": eb_id, "rate_date": rate_date},
        ).fetchone()

        record = {
            "row_index": r["row_index"],
            "emp_code": emp_code,
            "eb_id": eb_id,
            "rate": rate,
            "rate_date": rate_date.isoformat(),
        }
        if existing:
            record["existing_rate_id"] = int(existing.emp_rate_id)
            record["existing_rate"] = int(existing.rate) if existing.rate is not None else None
            existing_rows.append(record)
        else:
            new_rows.append(record)

    response_payload: dict = {
        "new_rows": new_rows,
        "existing_rows": existing_rows,
        "invalid_count": len(error_messages),
    }

    if error_messages:
        b64 = _build_invalid_workbook(wb, ws, list(error_messages.keys()), error_messages)
        original = file.filename or "rate_upload.xlsx"
        if original.lower().endswith(".xlsx"):
            invalid_name = original[:-5] + "_errors.xlsx"
        else:
            invalid_name = "rate_upload_errors.xlsx"
        response_payload["invalid_file"] = b64
        response_payload["invalid_filename"] = invalid_name

    return response_payload


@router.post("/emp_rate_excel_commit")
async def emp_rate_excel_commit(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Insert new rows and (optionally) overwrite existing ones.

    Body:
        {
          new_rows:      [{eb_id, rate, rate_date}],
          existing_rows: [{existing_rate_id?, eb_id, rate, rate_date}],
          overwrite:     bool   # if true, update existing_rows; else skip
        }
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    body = await request.json()
    new_rows = body.get("new_rows") or []
    existing_rows = body.get("existing_rows") or []
    overwrite = bool(body.get("overwrite", False))

    inserted = 0
    updated = 0
    skipped = 0

    try:
        now = datetime.now()

        for r in new_rows:
            eb_id = r.get("eb_id")
            rate = r.get("rate")
            rate_date = r.get("rate_date")
            if eb_id is None or rate is None or not rate_date:
                continue
            db.execute(
                text("""
                    INSERT INTO employee_rate_table (eb_id, rate, rate_date, updated_date_time)
                    VALUES (:eb_id, :rate, :rate_date, :ts)
                """),
                {
                    "eb_id": int(eb_id),
                    "rate": int(float(rate)),
                    "rate_date": rate_date,
                    "ts": now,
                },
            )
            inserted += 1

        if overwrite:
            for r in existing_rows:
                rate_id = r.get("existing_rate_id")
                eb_id = r.get("eb_id")
                rate = r.get("rate")
                rate_date = r.get("rate_date")
                if rate is None:
                    continue
                if rate_id:
                    db.execute(
                        text("""
                            UPDATE employee_rate_table
                            SET rate = :rate, updated_date_time = :ts
                            WHERE emp_rate_id = :rate_id
                        """),
                        {"rate": int(float(rate)), "ts": now, "rate_id": int(rate_id)},
                    )
                else:
                    db.execute(
                        text("""
                            UPDATE employee_rate_table
                            SET rate = :rate, updated_date_time = :ts
                            WHERE eb_id = :eb_id AND rate_date = :rate_date
                        """),
                        {
                            "rate": int(float(rate)),
                            "ts": now,
                            "eb_id": int(eb_id),
                            "rate_date": rate_date,
                        },
                    )
                updated += 1
        else:
            skipped = len(existing_rows)

        db.commit()
        return {
            "message": "Upload completed",
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/emp_rate_excel_upload")
async def emp_rate_excel_upload(
    request: Request,
    response: Response,
    file: UploadFile = File(...),
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Single-step Excel upload using a temp staging table for speed.

    Pipeline (all SQL, minimal Python loops):
      1. Ensure ``temp_employee_rate_table`` exists (created on first call).
      2. TRUNCATE the staging table.
      3. Bulk INSERT all parsed rows (emp_code + parsed rate / rate_date +
         row_index for error reporting) using a single executemany.
      4. UPDATE the staging table joining ``hrms_ed_official_details`` to fill
         ``eb_id`` for every valid emp_code.
      5. SELECT the rows with NULL eb_id / rate / rate_date — these are the
         invalid rows.
      6. If ANY invalid → return success=false with a red-marked Excel file.
         (Staging table left as-is; truncated on next upload.)
      7. Else → DELETE FROM employee_rate_table WHERE (eb_id, rate_date) IN
         (SELECT … FROM staging), then INSERT INTO employee_rate_table SELECT
         FROM staging in one statement. Truncate staging on success.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    rows, wb, ws, _col_idx = _read_excel_rows(file_bytes)
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found in Excel")

    try:
        # 1. Ensure staging table exists. Idempotent (IF NOT EXISTS).
        db.execute(text("""
            CREATE TABLE IF NOT EXISTS temp_employee_rate_table (
                row_index   INT          NOT NULL,
                emp_code    VARCHAR(50)  NULL,
                rate        INT          NULL,
                rate_date   DATE         NULL,
                eb_id       INT          NULL,
                INDEX idx_temp_emprate_eb_date (eb_id, rate_date)
            ) ENGINE=InnoDB
        """))

        # 2. Clear previous run.
        db.execute(text("TRUNCATE TABLE temp_employee_rate_table"))

        # 3. Bulk-load parsed rows. NULL out invalid rate / rate_date / emp_code
        #    so the validation SQL below can flag them uniformly.
        staging_rows = []
        for r in rows:
            emp_code = r["emp_code"] or None
            rate = _parse_rate(r["rate_raw"])
            rate_date = _parse_rate_date(r["date_raw"])
            staging_rows.append({
                "row_index": r["row_index"],
                "emp_code": emp_code,
                "rate": rate,
                "rate_date": rate_date,
            })

        # executemany via a single text() + list of dicts
        db.execute(
            text("""
                INSERT INTO temp_employee_rate_table
                    (row_index, emp_code, rate, rate_date)
                VALUES
                    (:row_index, :emp_code, :rate, :rate_date)
            """),
            staging_rows,
        )

        # 4. Resolve eb_id for valid emp_codes via a single UPDATE…JOIN.
        db.execute(text("""
            UPDATE temp_employee_rate_table t
            JOIN hrms_ed_official_details o
              ON o.emp_code = t.emp_code AND o.active = 1
            SET t.eb_id = o.eb_id
        """))

        # 5. Find invalid rows in one query.
        invalid_rows = db.execute(text("""
            SELECT row_index, emp_code, rate, rate_date, eb_id
            FROM temp_employee_rate_table
            WHERE eb_id IS NULL OR rate IS NULL OR rate_date IS NULL
        """)).fetchall()

        if invalid_rows:
            error_messages: dict[int, str] = {}
            for ir in invalid_rows:
                errs: list[str] = []
                if not ir.emp_code:
                    errs.append("emp_code is required")
                elif ir.eb_id is None:
                    errs.append("emp_code not found in employee master")
                if ir.rate is None:
                    errs.append("rate is invalid")
                if ir.rate_date is None:
                    errs.append("rate_date is invalid")
                error_messages[int(ir.row_index)] = "; ".join(errs) or "invalid row"

            # We didn't modify employee_rate_table, so no rollback needed for
            # data correctness; release the staging-table writes.
            db.commit()

            b64 = _build_invalid_workbook(
                wb, ws, list(error_messages.keys()), error_messages
            )
            original = file.filename or "rate_upload.xlsx"
            invalid_name = (
                original[:-5] + "_errors.xlsx"
                if original.lower().endswith(".xlsx")
                else "rate_upload_errors.xlsx"
            )
            return {
                "success": False,
                "message": "Upload failed",
                "invalid_count": len(error_messages),
                "invalid_file": b64,
                "invalid_filename": invalid_name,
            }

        # 6. All valid → bulk DELETE + bulk INSERT in two SQL statements.
        del_result = db.execute(text("""
            DELETE r FROM employee_rate_table r
            JOIN temp_employee_rate_table t
              ON t.eb_id = r.eb_id AND t.rate_date = r.rate_date
        """))
        deleted = int(del_result.rowcount or 0)

        ins_result = db.execute(text("""
            INSERT INTO employee_rate_table (eb_id, rate, rate_date, updated_date_time)
            SELECT eb_id, rate, rate_date, NOW()
            FROM temp_employee_rate_table
        """))
        inserted = int(ins_result.rowcount or 0)

        # 7. Clean up staging.
        db.execute(text("TRUNCATE TABLE temp_employee_rate_table"))

        db.commit()
        return {
            "success": True,
            "message": "Upload successful",
            "deleted": deleted,
            "inserted": inserted,
            "total": len(rows),
        }
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
