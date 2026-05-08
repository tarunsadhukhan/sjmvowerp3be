"""HRMS Bio Attendance bulk upload endpoints.

Flow:
  1. Client POSTs file to `/bio_att_upload`.
     - If `temp_bio_attendance_table` already has rows, the request is
       refused with "Another process Working".
     - Otherwise the file is parsed and bulk-inserted into the temp table
       *without any validation* (fast path).
     - A background job is queued that streams rows from the temp table,
       validates them against `bio_attendance_table` (duplicate check across
       the 10-column signature), inserts the survivors, and finally truncates
       the temp table.
  2. Client polls `/bio_att_excel_status/{job_id}` for progress.
  3. Client may call `/bio_att_clear` at any time to wipe the temp table.

Endpoints:
  - GET  /bio_att_list
  - POST /bio_att_upload
  - POST /bio_att_clear
  - GET  /bio_att_excel_status/{job_id}
"""

from __future__ import annotations

import csv
import io
import os
import traceback
import uuid
from datetime import datetime, date, time as dt_time, timedelta

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.sql import text

from src.authorization.utils import get_current_user_with_refresh
from src.config.db import extract_subdomain_from_request, get_engine, get_tenant_db

router = APIRouter()

BATCH_SIZE = 2000  # rows per INSERT into the temp table


# ---------------------------------------------------------------------------
# Job-status persistence (bio_att_jobs table — survives across uvicorn workers)
# ---------------------------------------------------------------------------

_JOB_FIELDS = (
    "status", "total", "processed", "inserted",
    "duplicates", "invalid", "message", "error",
)

_JOB_UPSERT_SQL = text(
    """
    INSERT INTO bio_att_jobs
        (job_id, status, total, processed, inserted, duplicates, invalid, message, error)
    VALUES
        (:job_id, :status, :total, :processed, :inserted, :duplicates, :invalid, :message, :error)
    ON DUPLICATE KEY UPDATE
        status     = COALESCE(VALUES(status), status),
        total      = COALESCE(VALUES(total), total),
        processed  = COALESCE(VALUES(processed), processed),
        inserted   = COALESCE(VALUES(inserted), inserted),
        duplicates = COALESCE(VALUES(duplicates), duplicates),
        invalid    = COALESCE(VALUES(invalid), invalid),
        message    = COALESCE(VALUES(message), message),
        error      = COALESCE(VALUES(error), error)
    """
)

_JOB_SELECT_SQL = text(
    """
    SELECT job_id, status, total, processed, inserted, duplicates, invalid,
           message, error, created_at, updated_at
    FROM bio_att_jobs WHERE job_id = :job_id
    """
)


def _set_job(db: Session, job_id: str, **fields) -> None:
    """UPSERT a row in bio_att_jobs. Unspecified fields are left untouched."""
    params: dict = {f: None for f in _JOB_FIELDS}
    params.update({k: v for k, v in fields.items() if k in _JOB_FIELDS})
    params["job_id"] = job_id
    db.execute(_JOB_UPSERT_SQL, params)
    db.commit()


def _get_job(db: Session, job_id: str) -> dict | None:
    row = db.execute(_JOB_SELECT_SQL, {"job_id": job_id}).fetchone()
    if not row:
        return None
    d = dict(row._mapping)
    for ts_field in ("created_at", "updated_at"):
        v = d.get(ts_field)
        if isinstance(v, datetime):
            d[ts_field] = v.strftime("%Y-%m-%d %H:%M:%S")
    return d


# Columns that compose the duplicate signature in bio_attendance_table
DUP_COLUMNS = [
    "emp_code",
    "emp_anme",
    "bio_id",
    "log_date",
    "company_name",
    "department",
    "designation",
    "employement_type",
    "device_direction",
    "device_name",
]

# Mapping accepted column headers (lowercased) -> table column name.
# Both the device-style CSV and the table-style headers are accepted.
HEADER_ALIASES = {
    "employee code": "emp_code",
    "emp_code": "emp_code",

    "employee name": "emp_anme",
    "emp_anme": "emp_anme",
    "emp_name": "emp_anme",

    "employee code in device": "bio_id",
    "bio_id": "bio_id",

    "logdate": "log_date",
    "log date": "log_date",
    "log_date": "log_date",

    "company": "company_name",
    "company_name": "company_name",

    "department": "department",
    "designation": "designation",

    "employement type": "employement_type",
    "employment type": "employement_type",
    "employement_type": "employement_type",

    "direction": "device_direction",
    "device_direction": "device_direction",

    "device name": "device_name",
    "device_name": "device_name",
}

IGNORED_HEADERS = {"category"}


# ---------------------------------------------------------------------------
# Listing
# ---------------------------------------------------------------------------


@router.get("/bio_att_list")
async def bio_att_list(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated bio-attendance listing."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))
        search = request.query_params.get("search")
        offset = max(page - 1, 0) * limit

        params: dict = {"limit": limit, "offset": offset}
        clauses: list[str] = []

        if search:
            clauses.append(
                "(emp_code LIKE :s OR emp_anme LIKE :s "
                " OR device_name LIKE :s OR department LIKE :s)"
            )
            params["s"] = f"%{search}%"

        # Per-column filters (whitelist -> safe column reference).
        # Frontend sends `f_<field>` matching DataGrid column field names.
        bio_filterable: dict[str, str] = {
            "emp_code":         "emp_code",
            "emp_anme":         "emp_anme",
            "bio_id":           "CAST(bio_id AS CHAR)",
            "log_date":         "CAST(log_date AS CHAR)",
            "department":       "department",
            "designation":      "designation",
            "device_direction": "device_direction",
            "device_name":      "device_name",
            "company_name":     "company_name",
            "employement_type": "employement_type",
        }
        for fld, expr in bio_filterable.items():
            v = request.query_params.get(f"f_{fld}")
            if v is None:
                continue
            v = v.strip()
            if not v:
                continue
            pname = f"f_{fld}"
            clauses.append(f"{expr} LIKE :{pname}")
            params[pname] = f"%{v}%"

        # Eb No (popup): exact emp_code match.
        emp_eq = (request.query_params.get("emp_code_eq") or "").strip()
        if emp_eq:
            clauses.append("emp_code = :emp_code_eq")
            params["emp_code_eq"] = emp_eq

        # Date-range filter: from_date / to_date applied to log_date.
        from_date = (request.query_params.get("from_date") or "").strip()
        to_date   = (request.query_params.get("to_date")   or "").strip()
        if from_date:
            clauses.append("DATE(log_date) >= :from_date")
            params["from_date"] = from_date
        if to_date:
            clauses.append("DATE(log_date) <= :to_date")
            params["to_date"] = to_date

        where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

        rows = db.execute(
            text(f"""
                SELECT bio_att_id, emp_code, emp_anme, bio_id, log_date,
                       company_name, department, designation,
                       employement_type, device_direction, device_name
                FROM bio_attendance_table
                {where}
                ORDER BY log_date DESC, bio_att_id DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        ).fetchall()

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        total = db.execute(
            text(f"SELECT COUNT(*) AS cnt FROM bio_attendance_table {where}"),
            count_params,
        ).fetchone()

        data = []
        for r in rows:
            d = dict(r._mapping)
            ld = d.get("log_date")
            if isinstance(ld, datetime):
                d["log_date"] = ld.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(ld, date):
                d["log_date"] = ld.isoformat()
            data.append(d)

        return {
            "data": data,
            "total": int(total.cnt) if total else 0,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/daily_att_list")
async def daily_att_list(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Paginated daily_attendance_process_table listing."""
    try:
        co_id = request.query_params.get("co_id")
        if not co_id:
            raise HTTPException(status_code=400, detail="co_id is required")

        page = int(request.query_params.get("page", 1))
        limit = int(request.query_params.get("limit", 10))
        search = request.query_params.get("search")
        offset = max(page - 1, 0) * limit

        params: dict = {"limit": limit, "offset": offset}
        clauses: list[str] = []

        if search:
            clauses.append(
                "(o.emp_code LIKE :s "
                " OR CONCAT_WS(' ', p.first_name, IFNULL(p.middle_name,''), IFNULL(p.last_name,'')) LIKE :s "
                " OR b.department LIKE :s "
                " OR b.designation LIKE :s "
                " OR d.spell_name LIKE :s "
                " OR d.attendance_type LIKE :s "
                " OR d.device_name LIKE :s "
                " OR CAST(d.attendance_date AS CHAR) LIKE :s)"
            )
            params["s"] = f"%{search}%"

        # Per-column filters (whitelist mapped to qualified SQL expressions).
        daily_filterable: dict[str, str] = {
            "daily_att_proc_id": "CAST(d.daily_att_proc_id AS CHAR)",
            "bio_id":            "CAST(d.bio_id AS CHAR)",
            "emp_code":          "o.emp_code",
            "emp_name":          "TRIM(CONCAT_WS(' ', p.first_name, IFNULL(p.middle_name,''), IFNULL(p.last_name,'')))",
            "department":        "b.department",
            "designation":       "b.designation",
            "attendance_date":   "CAST(d.attendance_date AS CHAR)",
            "spell_name":        "d.spell_name",
            "attendance_type":   "d.attendance_type",
            "attendance_source": "d.attendance_source",
            "check_in":          "CAST(d.check_in AS CHAR)",
            "check_out":         "CAST(d.check_out AS CHAR)",
            "Time_duration":     "CAST(d.Time_duration AS CHAR)",
            "Working_hours":     "CAST(d.Working_hours AS CHAR)",
            "Ot_hours":          "CAST(d.Ot_hours AS CHAR)",
            "device_name":       "d.device_name",
        }
        for fld, expr in daily_filterable.items():
            v = request.query_params.get(f"f_{fld}")
            if v is None:
                continue
            v = v.strip()
            if not v:
                continue
            pname = f"f_{fld}"
            clauses.append(f"{expr} LIKE :{pname}")
            params[pname] = f"%{v}%"

        # Eb No (popup): exact emp_code match.
        emp_eq = (request.query_params.get("emp_code_eq") or "").strip()
        if emp_eq:
            clauses.append("o.emp_code = :emp_code_eq")
            params["emp_code_eq"] = emp_eq

        # Date-range filter: from_date / to_date applied to attendance_date.
        from_date = (request.query_params.get("from_date") or "").strip()
        to_date   = (request.query_params.get("to_date")   or "").strip()
        if from_date:
            clauses.append("d.attendance_date >= :from_date")
            params["from_date"] = from_date
        if to_date:
            clauses.append("d.attendance_date <= :to_date")
            params["to_date"] = to_date

        where = (" WHERE " + " AND ".join(clauses)) if clauses else ""

        rows = db.execute(
            text(f"""
                SELECT d.daily_att_proc_id, d.bio_id,
                       o.emp_code AS emp_code,
                       TRIM(CONCAT_WS(' ', p.first_name,
                                          IFNULL(p.middle_name,''),
                                          IFNULL(p.last_name,''))) AS emp_name,
                       b.department AS department,
                       b.designation AS designation,
                       d.attendance_date, d.spell_name, d.attendance_type,
                       d.attendance_source, d.check_in, d.check_out,
                       d.Time_duration, d.Working_hours, d.Ot_hours,
                       d.spell_start_time, d.spell_end_time, d.spell_hours,
                       d.processed, d.device_name
                FROM daily_attendance_process_table d
                LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
                LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
                LEFT JOIN bio_attendance_table b     ON b.bio_att_id = d.bio_id
                {where}
                ORDER BY d.attendance_date DESC, d.daily_att_proc_id DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        ).fetchall()

        count_params = {k: v for k, v in params.items() if k not in ("limit", "offset")}
        total = db.execute(
            text(
                f"""SELECT COUNT(*) AS cnt
                    FROM daily_attendance_process_table d
                    LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
                    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
                    LEFT JOIN bio_attendance_table b     ON b.bio_att_id = d.bio_id
                    {where}"""
            ),
            count_params,
        ).fetchone()

        def _fmt(v):
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(v, date):
                return v.isoformat()
            if isinstance(v, timedelta):
                # MySQL TIME columns come back as timedelta in PyMySQL.
                total_seconds = int(v.total_seconds())
                h, rem = divmod(total_seconds, 3600)
                m, s = divmod(rem, 60)
                return f"{h:02d}:{m:02d}:{s:02d}"
            return v

        data = [{k: _fmt(v) for k, v in r._mapping.items()} for r in rows]

        return {
            "data": data,
            "total": int(total.cnt) if total else 0,
            "page": page,
            "limit": limit,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_log_date(value):
    """Parse a cell value into a datetime. Returns None if invalid."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    fmts = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y",
    )
    for fmt in fmts:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _to_int_or_none(value):
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _str_or_empty(value):
    return "" if value is None else str(value).strip()


def _resolve_header(raw):
    if raw is None:
        return None
    key = str(raw).strip().lower()
    if key in IGNORED_HEADERS:
        return None
    return HEADER_ALIASES.get(key)


def _read_upload(file_bytes: bytes, filename: str) -> list[dict]:
    """Return parsed rows from CSV / xlsx upload. Only structural validation
    (required columns must be present); no row-level validation is done here."""
    try:
        import openpyxl
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    is_csv = (filename or "").lower().endswith(".csv")

    if is_csv:
        try:
            text_content = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text_content = file_bytes.decode("latin-1")
        reader = csv.reader(io.StringIO(text_content))
        wb = openpyxl.Workbook()
        ws = wb.active
        for csv_row in reader:
            ws.append(csv_row)
    else:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=False)
    try:
        header_cells = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="File is empty")

    col_idx: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        db_col = _resolve_header(cell.value)
        if db_col and db_col not in col_idx:
            col_idx[db_col] = idx

    missing = [c for c in DUP_COLUMNS if c not in col_idx]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}",
        )

    rows: list[dict] = []
    for row_cells in rows_iter:
        values = [c.value for c in row_cells]
        if all((v is None or (isinstance(v, str) and not v.strip())) for v in values):
            continue

        def _val(name: str):
            i = col_idx[name]
            return values[i] if i < len(values) else None

        rows.append({
            "emp_code": _str_or_empty(_val("emp_code")) or None,
            "emp_anme": _str_or_empty(_val("emp_anme")) or None,
            "bio_id": _to_int_or_none(_val("bio_id")),
            "log_date": _parse_log_date(_val("log_date")),
            "company_name": _str_or_empty(_val("company_name")) or None,
            "department": _str_or_empty(_val("department")) or None,
            "designation": _str_or_empty(_val("designation")) or None,
            "employement_type": _str_or_empty(_val("employement_type")) or None,
            "device_direction": _str_or_empty(_val("device_direction")) or None,
            "device_name": _str_or_empty(_val("device_name")) or None,
        })
    return rows


# ---------------------------------------------------------------------------
# SQL constants
# ---------------------------------------------------------------------------

TEMP_INSERT_SQL = text(
    """
    INSERT INTO temp_bio_attendance_table
        (emp_code, emp_anme, bio_id, log_date, company_name,
         department, designation, employement_type,
         device_direction, device_name)
    VALUES
        (:emp_code, :emp_anme, :bio_id, :log_date, :company_name,
         :department, :designation, :employement_type,
         :device_direction, :device_name)
    """
)

TEMP_COUNT_SQL = text("SELECT COUNT(*) AS cnt FROM temp_bio_attendance_table")
TEMP_DELETE_SQL = text("DELETE FROM temp_bio_attendance_table")

# Only rows captured from these devices are eligible to move into
# bio_attendance_table. Anything else is treated as invalid/ignored.
ALLOWED_DEVICE_NAMES = ("AIFace(Mars) Out-F", "AIFace(Mars) In-F")

# Count rows that would be rejected as invalid (missing emp_code or log_date,
# or device_name not in the allow-list).
INVALID_COUNT_SQL = text(
    """
    SELECT COUNT(*) AS cnt
    FROM temp_bio_attendance_table
    WHERE emp_code IS NULL OR emp_code = '' OR log_date IS NULL
       OR device_name IS NULL
       OR device_name NOT IN ('AIFace(Mars) Out-F', 'AIFace(Mars) In-F')
    """
)

# Set-difference move: temp MINUS bio (NULL-safe on all 10 cols).
# Equivalent to:
#   SELECT * FROM temp
#   MINUS
#   SELECT * FROM bio
# but expressed with NOT EXISTS so MySQL can use indexes.
BULK_INSERT_SQL = text(
    """
    INSERT INTO bio_attendance_table
        (emp_code, emp_anme, bio_id, log_date, company_name,
         department, designation, employement_type,
         device_direction, device_name)
    SELECT t.emp_code, t.emp_anme, t.bio_id, t.log_date, t.company_name,
           t.department, t.designation, t.employement_type,
           t.device_direction, t.device_name
    FROM temp_bio_attendance_table t
    WHERE t.emp_code IS NOT NULL AND t.emp_code <> ''
      AND t.log_date IS NOT NULL
      AND t.device_name IN ('AIFace(Mars) Out-F', 'AIFace(Mars) In-F')
      AND NOT EXISTS (
          SELECT 1 FROM bio_attendance_table b
          WHERE b.emp_code = t.emp_code
            AND b.log_date = t.log_date
            AND b.emp_anme <=> t.emp_anme
            AND b.bio_id <=> t.bio_id
            AND b.company_name <=> t.company_name
            AND b.department <=> t.department
            AND b.designation <=> t.designation
            AND b.employement_type <=> t.employement_type
            AND b.device_direction <=> t.device_direction
            AND b.device_name <=> t.device_name
      )
    """
)


# ---------------------------------------------------------------------------
# Link-master back-fill (runs after BULK_INSERT_SQL).
#
# tbl_master_bio_link_mst columns (per user):
#   match_type   CHAR(1)        -- 'E', 'D', 'O', 'B'
#   bio_data     VARCHAR(...)   -- raw text as it appears in bio_attendance_table
#   master_id    INT            -- internal id from sjm masters
#
# Mapping:
#   E : bio_data = bio_attendance_table.emp_code     -> eb_id
#   D : bio_data = bio_attendance_table.department   -> dept_id
#   O : bio_data = bio_attendance_table.designation  -> desig_id
#   B : bio_data = bio_attendance_table.device_name  -> device_id
# ---------------------------------------------------------------------------

# (match_type, bio_attendance target column, bio_attendance source column)
_LINK_UPDATE_TEMPLATES = (
    ("E", "eb_id",     "emp_code"),
    ("D", "dept_id",   "department"),
    ("O", "desig_id",  "designation"),
    ("B", "device_id", "device_name"),
)


def _backfill_links(db: Session) -> dict:
    """Populate eb_id / dept_id / desig_id / device_id on bio_attendance_table
    using tbl_master_bio_link_mst (columns: match_type, bio_data, master_id).
    Returns rowcounts per match_type. Each UPDATE is wrapped so a failure
    on one (e.g. missing column on bio_attendance_table) doesn't abort
    the others."""
    counts: dict = {"E": 0, "D": 0, "O": 0, "B": 0, "errors": []}
    for code, target_col, match_col in _LINK_UPDATE_TEMPLATES:
        sql = text(
            f"""
            UPDATE bio_attendance_table b
            JOIN tbl_master_bio_link_mst m
                ON m.match_type = :code
               AND m.bio_data = b.{match_col}
            SET b.{target_col} = m.master_id
            """
        )
        try:
            res = db.execute(sql, {"code": code})
            counts[code] = int(res.rowcount or 0)
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            counts["errors"].append(f"{code}: {type(e).__name__}: {e}")
    return counts


# ---------------------------------------------------------------------------
# Background validate-and-move worker
# ---------------------------------------------------------------------------


def _make_session(subdomain: str) -> Session:
    tenant_url = (
        f"mysql+pymysql://{os.getenv('DATABASE_USER')}:{os.getenv('DATABASE_PASSWORD')}"
        f"@{os.getenv('DATABASE_HOST')}:{os.getenv('DATABASE_PORT')}/{subdomain}"
    )
    engine = get_engine(tenant_url)
    SessionTenant = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    return SessionTenant()


def _run_validate_and_move(job_id: str, subdomain: str) -> None:
    """Set-based move: ``temp MINUS bio`` -> ``bio_attendance_table``.
    The temp table is **not** deleted afterwards (use /bio_att_clear).
    Status is persisted to ``bio_att_jobs`` so any uvicorn worker can read it.
    """
    db = _make_session(subdomain)

    try:
        total = int((db.execute(TEMP_COUNT_SQL).fetchone() or {"cnt": 0}).cnt)
        _set_job(
            db, job_id,
            status="running", total=total, processed=0,
            inserted=0, duplicates=0, invalid=0,
        )

        invalid = int((db.execute(INVALID_COUNT_SQL).fetchone() or {"cnt": 0}).cnt)

        # Single statement: insert rows that exist in temp but not in bio.
        result = db.execute(BULK_INSERT_SQL)
        inserted = int(result.rowcount or 0)
        db.commit()

        duplicates = max(total - invalid - inserted, 0)

        # Back-fill eb_id / dept_id / desig_id / device_id from the
        # link master. Best-effort — failures don't abort the upload.
        try:
            link_counts = _backfill_links(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            link_counts = {"errors": ["backfill failed"]}

        # Wipe the staging table now that the move has been committed.
        try:
            db.execute(TEMP_DELETE_SQL)
            db.commit()
        except Exception:
            db.rollback()

        link_msg = (
            f" Linked: E={link_counts.get('E', 0)}, D={link_counts.get('D', 0)}, "
            f"O={link_counts.get('O', 0)}, B={link_counts.get('B', 0)}."
        )
        if link_counts.get("errors"):
            link_msg += f" Link errors: {'; '.join(link_counts['errors'])}"

        _set_job(
            db, job_id,
            status="completed",
            total=total,
            processed=total,
            inserted=inserted,
            duplicates=duplicates,
            invalid=invalid,
            message="Upload completed." + link_msg,
        )
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        try:
            _set_job(
                db, job_id,
                status="failed",
                error=f"{e}\n{traceback.format_exc()}"[:60000],
            )
        except Exception:
            pass
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@router.post("/bio_att_upload")
async def bio_att_upload(
    request: Request,
    response: Response,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Bulk-load the uploaded file into the temp table, then queue a
    background job that validates and moves the data into the main table."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    # Refuse if a previous batch is still parked in the temp table.
    existing = db.execute(TEMP_COUNT_SQL).fetchone()
    if existing and int(existing.cnt) > 0:
        raise HTTPException(status_code=409, detail="Another process Working")

    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Empty file")

    rows = _read_upload(file_bytes, file.filename or "")
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows found")

    # Bulk insert into temp table in batches (no row-level validation).
    try:
        for start in range(0, len(rows), BATCH_SIZE):
            batch = rows[start : start + BATCH_SIZE]
            db.execute(TEMP_INSERT_SQL, batch)
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed loading temp table: {e}")

    subdomain = extract_subdomain_from_request(request)
    job_id = uuid.uuid4().hex
    _set_job(
        db, job_id,
        status="queued",
        total=len(rows),
        processed=0,
        inserted=0,
        duplicates=0,
        invalid=0,
    )
    background_tasks.add_task(_run_validate_and_move, job_id, subdomain)

    return {
        "message": "Upload accepted",
        "job_id": job_id,
        "queued": len(rows),
    }


@router.post("/bio_att_clear")
async def bio_att_clear(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Wipe the temp table. Manual override for stuck batches."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        existing = db.execute(TEMP_COUNT_SQL).fetchone()
        deleted = int(existing.cnt) if existing else 0
        db.execute(TEMP_DELETE_SQL)
        db.commit()
        return {"message": "Temp table cleared", "deleted": deleted}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/bio_att_excel_status/{job_id}")
async def bio_att_excel_status(
    job_id: str,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    job = _get_job(db, job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return {"job_id": job_id, **job}


@router.get("/bio_att_temp_count")
async def bio_att_temp_count(
    request: Request,
    response: Response,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Return current row count of temp_bio_attendance_table."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")
    try:
        row = db.execute(TEMP_COUNT_SQL).fetchone()
        return {"count": int(row.cnt) if row else 0}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Bio-attendance -> daily_attendance_process_table  (Process button)
# =============================================================================
#
# Algorithm (per emp x date):
#   * Validate emp_code against hrms_ed_official_details (active=1) -> eb_id.
#       Unmatched emp_codes for the date are returned as an XLSX download
#       and the entire process is aborted.
#   * Spell A:  IN-time window 05:30..09:59:59,  work window 06:00..14:00
#   * Spell B:  IN-time window 13:30..15:59:59,  work window 14:00..22:00
#       (No Spell C / night shift.)
#   * Working day vs weekly off:
#       - tbl_offday_mst.day_off stores int 0..6 (Sun..Sat).
#       - If the attendance_date weekday matches a day_off row -> off day.
#       - Off day spell row: attendance_type = 'O'
#       - Working day spell row: attendance_type = 'P'
#   * Overtime split:
#       - Only on working days.
#       - If raw duration > 8.5h, additionally insert one row with
#         spell_name = same A/B, attendance_type = 'O',
#         Working_hours = duration - 8.0,  Ot_hours = duration - 8.0.
#   * Re-run on same date = DELETE existing rows for that date, re-insert.

SPELL_A_IN_FROM = "05:30:00"
SPELL_A_IN_TO   = "10:00:00"  # exclusive
SPELL_A_START   = "06:00:00"
SPELL_A_END     = "14:00:00"

SPELL_B_IN_FROM = "13:30:00"
SPELL_B_IN_TO   = "16:00:00"  # exclusive
SPELL_B_START   = "14:00:00"
SPELL_B_END     = "22:00:00"

SPELL_HOURS = 8.0
OT_THRESHOLD_HOURS = 8.5


def _log_sql(label: str, stmt, params: dict | None = None) -> None:
    """Print a SQL statement + bind params to stdout (uvicorn console).
    Used for live debugging the Process endpoint."""
    try:
        sql_str = str(getattr(stmt, "text", stmt))
    except Exception:
        sql_str = repr(stmt)
    print(f"\n[bio_att_process] === {label} ===", flush=True)
    print(sql_str.strip(), flush=True)
    if params is not None:
        print(f"[bio_att_process] params = {params}", flush=True)


# 1. Find rows on :tran_date that are NOT fully linked
#    (any of eb_id / dept_id / desig_id / device_id is NULL).
#    These are reported as the "unmatched" xlsx and skipped during processing.
UNMATCHED_EMP_CODES_SQL = text(
    """
    SELECT b.emp_code, b.emp_anme, b.bio_id, b.log_date,
           b.company_name, b.department, b.designation,
           b.employement_type, b.device_direction, b.device_name
    FROM bio_attendance_table b
    WHERE DATE(b.log_date) = :tran_date
      AND (b.eb_id IS NULL
        OR b.dept_id IS NULL
        OR b.desig_id IS NULL
        OR b.device_id IS NULL)
    ORDER BY b.emp_code, b.log_date
    """
)

# 2. Did weekday match a day_off row? (Returns count > 0 if it's a weekly-off day.)
#    Convention: tbl_offday_mst.off_day uses 1=Sun, 2=Mon, ..., 7=Sat.
#    MySQL DAYOFWEEK() also returns 1=Sun..7=Sat, so use it directly.
IS_OFF_DAY_SQL = text(
    "SELECT COUNT(*) AS cnt FROM tbl_offday_mst WHERE off_day = DAYOFWEEK(:tran_date)"
)

# 3. Wipe re-run rows for the date.
DELETE_DAY_ROWS_SQL = text(
    "DELETE FROM daily_attendance_process_table WHERE attendance_date = :tran_date"
)

# 4. Fetch all punches for the spell window. We then pair them in Python
#    (1st=IN, 2nd=OUT, 3rd=IN, 4th=OUT, ...) and only count pairs whose
#    duration is at least 1 hour as Working_hours. check_in is the first
#    punch, check_out is the last punch.
#
#    Only rows where ALL four link ids (eb_id, dept_id, desig_id, device_id)
#    are populated are considered. Window: punches between :in_from and
#    :spell_end (inclusive of in_from, inclusive of spell_end+OT slack).
FETCH_SPELL_PUNCHES_SQL = text(
    """
    SELECT b.eb_id,
           b.bio_att_log_id,
           b.dept_id,
           b.desig_id,
           TIME(b.log_date)   AS punch_time,
           b.log_date         AS log_date
    FROM bio_attendance_table b
    WHERE DATE(b.log_date) = :tran_date
      AND b.eb_id     IS NOT NULL
      AND b.dept_id   IS NOT NULL
      AND b.desig_id  IS NOT NULL
      AND b.device_id IS NOT NULL
      AND TIME(b.log_date) >= :in_from
      AND TIME(b.log_date) <= :window_end
    ORDER BY b.eb_id, b.log_date
    """
)

# Insert one daily row per employee using bound params.
INSERT_SPELL_ROW_SQL = text(
    """
    INSERT INTO daily_attendance_process_table
        (eb_id, bio_id, dept_id, desig_id,
         attendance_date, spell_name, attendance_type,
         attendance_source, check_in, check_out,
         Time_duration, Working_hours, Ot_hours,
         spell_start_time, spell_end_time, spell_hours, processed)
    VALUES
        (:eb_id, :bio_id, :dept_id, :desig_id,
         :tran_date, :spell_name, :attendance_type,
         'BIO', :check_in, :check_out,
         :time_duration, :working_hours, :ot_hours,
         :spell_start, :spell_end, :spell_hours, 1)
    """
)

# Minimum pair duration that counts toward Working_hours.
MIN_PAIR_SECONDS = 3600  # 1 hour


def _bucket_hours(h: float) -> float:
    """Bucket hours: >=7 -> 8, [3, 7) -> 4, <3 -> 0."""
    if h >= 7:
        return 8.0
    if h >= 3:
        return 4.0
    return 0.0


def _parse_hms(s: str) -> dt_time:
    """Parse 'HH:MM:SS' into datetime.time."""
    h, m, sec = (int(x) for x in s.split(":"))
    return dt_time(h, m, sec)


def _to_seconds(t) -> int:
    """Convert a datetime.time / timedelta-ish value to seconds-of-day."""
    if isinstance(t, timedelta):
        return int(t.total_seconds())
    if isinstance(t, dt_time):
        return t.hour * 3600 + t.minute * 60 + t.second
    # Fallback: parse string 'HH:MM:SS'
    return _to_seconds(_parse_hms(str(t)))


def _process_one_spell(
    db: Session,
    *,
    tran_date: str,
    spell_name: str,
    in_from: str,
    in_to: str,
    spell_start: str,
    spell_end: str,
    is_off_day: bool,
) -> tuple[int, int]:
    """Build daily_attendance_process_table rows for this spell.

    Logic:
      * Pull every punch (per emp) in [in_from .. spell_end+slack] for the date.
      * The employee belongs to this spell iff their first punch is within
        [in_from .. in_to).
      * Total duration = (last_punch - first_punch). Skip if < 1 hour.
      * Working_hours_raw = min(duration, 8); ot_hours_raw = max(0, duration - 8).
      * Bucket each: >=7 -> 8, [3, 7) -> 4, <3 -> 0.
      * Time_duration = working_hours + ot_hours (post-bucket).
      * On an off day, ot_hours = 0.

    Returns (regular_inserted, 0). OT does not create a separate row.
    """
    in_from_sec = _to_seconds(_parse_hms(in_from))
    in_to_sec   = _to_seconds(_parse_hms(in_to))

    # Slack so that a late check-out after spell_end is still captured.
    # Use a wide window (effectively end-of-day) so OT punches several
    # hours past spell_end aren't dropped. The spell that an employee
    # belongs to is decided below by their FIRST punch falling within
    # [in_from .. in_to), so widening the upper bound is safe.
    OT_SLACK_HOURS = 16
    spell_end_t = _parse_hms(spell_end)
    window_end_dt = (datetime.combine(date.min, spell_end_t)
                     + timedelta(hours=OT_SLACK_HOURS))
    # cap at 23:59:59 (we don't cross midnight here)
    if window_end_dt.day != date.min.day:
        window_end = "23:59:59"
    else:
        window_end = window_end_dt.time().strftime("%H:%M:%S")

    fetch_params = {
        "tran_date": tran_date,
        "in_from": in_from,
        "window_end": window_end,
    }
    _log_sql(f"FETCH_SPELL_PUNCHES_SQL [spell={spell_name}]",
             FETCH_SPELL_PUNCHES_SQL, fetch_params)
    rows = db.execute(FETCH_SPELL_PUNCHES_SQL, fetch_params).fetchall()
    print(f"[bio_att_process]   -> fetched {len(rows)} punch row(s)", flush=True)

    # Group punches per employee, preserving SQL ORDER BY (eb_id, log_date).
    # Each tuple: (punch_time, bio_att_log_id, dept_id, desig_id, log_date).
    by_emp: dict[int, list[tuple]] = {}
    for r in rows:
        m = r._mapping
        by_emp.setdefault(m["eb_id"], []).append(
            (m["punch_time"], m["bio_att_log_id"], m["dept_id"], m["desig_id"], m["log_date"])
        )

    inserted_reg = 0
    inserted_ot  = 0
    base_attendance_type = "O" if is_off_day else "P"

    for eb_id, punches in by_emp.items():
        if not punches:
            continue
        first_time, first_bio, first_dept, first_desig, first_log_date = punches[0]
        first_sec = _to_seconds(first_time)
        # Employee only belongs to this spell if first punch is inside IN window.
        if not (in_from_sec <= first_sec < in_to_sec):
            continue

        last_time, _, _, _, last_log_date = punches[-1]
        last_sec = _to_seconds(last_time)

        # Total span: last punch - first punch (covers etrack data where
        # only IN/OUT are recorded, and gives the "full presence" duration).
        span_secs = max(0, last_sec - first_sec)
        if span_secs < MIN_PAIR_SECONDS:
            # Less than 1 hour on site -- skip row.
            continue

        total_hours    = round(span_secs / 3600.0, 2)
        # Working hours = duration capped at 8; OT = anything beyond 8.
        capped_working = round(min(SPELL_HOURS, total_hours), 2)
        if is_off_day:
            ot_raw = 0.0
        else:
            ot_raw = round(max(total_hours - SPELL_HOURS, 0.0), 2)
        # Bucket both: >=7 -> 8, [3, 7) -> 4, <3 -> 0.
        capped_working = _bucket_hours(capped_working)
        ot_hours = _bucket_hours(ot_raw)
        # Time_duration = total of bucketed working + OT.
        time_duration = round(capped_working + ot_hours, 2)

        reg_params = {
            "eb_id": int(eb_id),
            "bio_id": int(first_bio) if first_bio is not None else None,
            "dept_id": int(first_dept) if first_dept is not None else None,
            "desig_id": int(first_desig) if first_desig is not None else None,
            "tran_date": tran_date,
            "spell_name": spell_name,
            "attendance_type": base_attendance_type,
            "check_in": first_log_date,
            "check_out": last_log_date,
            "time_duration": time_duration,
            "working_hours": capped_working,
            "ot_hours": ot_hours,
            "spell_start": spell_start,
            "spell_end": spell_end,
            "spell_hours": SPELL_HOURS,
        }
        _log_sql(f"INSERT_SPELL_ROW_SQL [spell={spell_name} eb_id={eb_id}]",
                 INSERT_SPELL_ROW_SQL, reg_params)
        db.execute(INSERT_SPELL_ROW_SQL, reg_params)
        inserted_reg += 1

    inserted_ot = 0  # OT no longer creates a separate row.
    print(f"[bio_att_process]   -> rows inserted = {inserted_reg}", flush=True)
    return inserted_reg, inserted_ot


# ─────────────────────────────────────────────────────────────────────────────
# Etrack-specific processing. Used by /bio_att_etrack_process.
# May produce ONE OR TWO daily_attendance_process_table rows per (eb_id, date)
# depending on the first-entry window and last-entry time.
#
# Notation (decimal hours-of-day):
#   first_h = first_punch hour-of-day      e.g. 5:30 -> 5.5
#   last_h  = last_punch  hour-of-day      e.g. 17:00 -> 17.0
#   span    = last_h - first_h             (must be >= 1 hour to qualify)
#
# Rule 1 — first_h in [5, 9):
#   eff = last_h - 6
#   if eff >= 7        -> rec1 {spell:A,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:A1, type:R, working:4}
#   O = eff - working_hours_of_rec1
#   if O >= 7          -> rec2 {spell:B,  type:O, ot:8}
#   elif O in [3,7)    -> rec2 {spell:B1, type:O, ot:4}
#
# Rule 2 — first_h in [9, 13):
#   if span >= 3       -> rec1 {spell:A2, type:O, ot:4}
#   if span > 4:
#     d1 = last_h - 14
#     if d1 >= 7       -> rec2 {spell:B,  type:R, working:8}
#     elif d1 in [3,7) -> rec2 {spell:B1, type:R, working:4}
#
# Rule 3 — first_h in [13, 21):
#   eff = last_h - 14
#   if eff >= 7        -> rec1 {spell:B,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:B1, type:R, working:4}
#
# On an off day, the spell + working/ot calculations are unchanged; only the
# attendance_type is forced to "O" on every row.
# ─────────────────────────────────────────────────────────────────────────────

_ETRACK_SPELL_TIMES: dict[str, tuple[str, str]] = {
    "A":  ("06:00:00", "14:00:00"),
    "A1": ("06:00:00", "14:00:00"),
    "A2": ("09:00:00", "13:00:00"),
    "B":  ("14:00:00", "22:00:00"),
    "B1": ("14:00:00", "22:00:00"),
}


def _etrack_records_for_employee(
    first_h: float, last_h: float, span: float,
) -> list[dict]:
    """Apply rules #1/#2/#3 and return a list of {spell, att_type, working, ot}
    dicts (0..2 entries)."""
    out: list[dict] = []

    # Rule 1: first entry [5, 9)
    if 5 <= first_h < 9:
        eff = last_h - 6
        rec1_working = 0.0
        if eff >= 7:
            out.append({"spell": "A",  "att_type": "R", "working": 8.0, "ot": 0.0})
            rec1_working = 8.0
        elif eff >= 3:
            out.append({"spell": "A1", "att_type": "R", "working": 4.0, "ot": 0.0})
            rec1_working = 4.0
        if rec1_working > 0:
            O = eff - rec1_working
            if O >= 7:
                out.append({"spell": "B",  "att_type": "O", "working": 0.0, "ot": 8.0})
            elif O >= 3:
                out.append({"spell": "B1", "att_type": "O", "working": 0.0, "ot": 4.0})

    # Rule 2: first entry [9, 13)
    elif 9 <= first_h < 13:
        if span >= 3:
            out.append({"spell": "A2", "att_type": "O", "working": 0.0, "ot": 4.0})
        if span > 4:
            d1 = last_h - 14
            if d1 >= 7:
                out.append({"spell": "B",  "att_type": "R", "working": 8.0, "ot": 0.0})
            elif d1 >= 3:
                out.append({"spell": "B1", "att_type": "R", "working": 4.0, "ot": 0.0})

    # Rule 3: first entry [13, 21)
    elif 13 <= first_h < 21:
        eff = last_h - 14
        if eff >= 7:
            out.append({"spell": "B",  "att_type": "R", "working": 8.0, "ot": 0.0})
        elif eff >= 3:
            out.append({"spell": "B1", "att_type": "R", "working": 4.0, "ot": 0.0})

    return out


def _process_etrack_day(
    db: Session,
    *,
    tran_date: str,
    is_off_day: bool,
) -> int:
    """Insert daily_attendance_process_table rows for the day per the etrack
    rules. Returns total number of rows inserted (across all employees)."""
    rows = db.execute(
        FETCH_SPELL_PUNCHES_SQL,
        {"tran_date": tran_date, "in_from": "00:00:00", "window_end": "23:59:59"},
    ).fetchall()

    by_emp: dict[int, list[tuple]] = {}
    for r in rows:
        m = r._mapping
        by_emp.setdefault(m["eb_id"], []).append(
            (m["punch_time"], m["bio_att_log_id"], m["dept_id"], m["desig_id"], m["log_date"])
        )

    inserted = 0
    for eb_id, punches in by_emp.items():
        if len(punches) < 2:
            continue
        first_time, first_bio, first_dept, first_desig, first_log_date = punches[0]
        last_time, _, _, _, last_log_date = punches[-1]
        first_sec = _to_seconds(first_time)
        last_sec  = _to_seconds(last_time)
        span_secs = max(0, last_sec - first_sec)
        if span_secs < MIN_PAIR_SECONDS:
            continue

        first_h = first_sec / 3600.0
        last_h  = last_sec  / 3600.0
        span    = round(last_h - first_h, 4)

        recs = _etrack_records_for_employee(first_h, last_h, span)
        if not recs:
            continue

        for rec in recs:
            # On an off day, spell + working/ot stay the same, only the
            # attendance_type is forced to "O".
            att_type      = "O" if is_off_day else rec["att_type"]
            working_hours = rec["working"]
            ot_hours      = rec["ot"]
            time_duration = round(working_hours + ot_hours, 2)
            spell_start, spell_end = _ETRACK_SPELL_TIMES.get(
                rec["spell"], ("00:00:00", "00:00:00"),
            )
            db.execute(
                INSERT_SPELL_ROW_SQL,
                {
                    "eb_id": int(eb_id),
                    "bio_id": int(first_bio) if first_bio is not None else None,
                    "dept_id": int(first_dept) if first_dept is not None else None,
                    "desig_id": int(first_desig) if first_desig is not None else None,
                    "tran_date": tran_date,
                    "spell_name": rec["spell"],
                    "attendance_type": att_type,
                    "check_in": first_log_date,
                    "check_out": last_log_date,
                    "time_duration": time_duration,
                    "working_hours": working_hours,
                    "ot_hours": ot_hours,
                    "spell_start": spell_start,
                    "spell_end": spell_end,
                    "spell_hours": SPELL_HOURS,
                },
            )
            inserted += 1

    print(f"[bio_att_etrack_process]   -> rows inserted = {inserted}", flush=True)
    return inserted


def _build_unmatched_xlsx(rows) -> bytes:
    """Build an xlsx workbook listing the raw bio rows for unmatched emp_codes."""
    try:
        import openpyxl
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl not installed: {e}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unmatched Emp Codes"
    headers = [
        "emp_code", "emp_anme", "bio_id", "log_date", "company_name",
        "department", "designation", "employement_type",
        "device_direction", "device_name",
    ]
    ws.append(headers)
    for r in rows:
        m = r._mapping
        ws.append([
            m.get("emp_code"),
            m.get("emp_anme"),
            m.get("bio_id"),
            str(m.get("log_date") or ""),
            m.get("company_name"),
            m.get("department"),
            m.get("designation"),
            m.get("employement_type"),
            m.get("device_direction"),
            m.get("device_name"),
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.post("/bio_att_process")
async def bio_att_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Process raw bio_attendance_table rows for a date into
    daily_attendance_process_table.

    Request body: ``{"tran_date": "YYYY-MM-DD"}``.

    Behaviour:
      * If any emp_code on that date is unknown to ``hrms_ed_official_details``,
        responds with an XLSX attachment listing those raw rows and aborts —
        no inserts are made.
      * Otherwise wipes existing rows for that date and re-inserts spell A,
        spell B, and OT rows. Returns JSON stats.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body = await request.json()
    except Exception:
        body = {}
    tran_date = (body or {}).get("tran_date")
    if not tran_date:
        raise HTTPException(status_code=400, detail="tran_date is required")
    try:
        # Validate format.
        datetime.strptime(tran_date, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="tran_date must be YYYY-MM-DD")

    try:
        # Step 1: collect emp_codes that punched on this date but aren't in the
        # employee master. We don't abort — they're just skipped (the INSERT
        # JOINs against hrms_ed_official_details, so unmatched rows fall out
        # naturally). The list is returned as an xlsx attachment.
        try:
            _log_sql("UNMATCHED_EMP_CODES_SQL", UNMATCHED_EMP_CODES_SQL, {"tran_date": tran_date})
            unmatched = db.execute(
                UNMATCHED_EMP_CODES_SQL, {"tran_date": tran_date}
            ).fetchall()
            print(f"[bio_att_process] unmatched rows: {len(unmatched)}", flush=True)
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=(
                    "Failed reading bio_attendance_table / hrms_ed_official_details. "
                    f"Original error: {e}"
                ),
            )
        unique_codes = {r._mapping.get("emp_code") for r in unmatched}

        # Step 2: weekly-off check. Fail-soft — if tbl_offday_mst doesn't
        # exist or has unexpected schema, treat the day as a working day.
        try:
            _log_sql("IS_OFF_DAY_SQL", IS_OFF_DAY_SQL, {"tran_date": tran_date})
            off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
            is_off_day = bool(off_row and int(off_row.cnt) > 0)
            print(f"[bio_att_process] is_off_day = {is_off_day}", flush=True)
        except Exception as e:
            print(f"[bio_att_process] IS_OFF_DAY_SQL failed (treating as working day): {e}", flush=True)
            try:
                db.rollback()
            except Exception:
                pass
            is_off_day = False

        # Step 3: wipe existing rows for this date (re-run safe).
        try:
            _log_sql("DELETE_DAY_ROWS_SQL", DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
            db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
        except Exception as e:
            traceback.print_exc()
            raise HTTPException(
                status_code=500,
                detail=(
                    "daily_attendance_process_table not accessible. "
                    "Create the table (see DDL in repo notes). "
                    f"Original error: {e}"
                ),
            )

        # Step 4: process Spell A and Spell B for matched employees only.
        try:
            a_reg, a_ot = _process_one_spell(
                db, tran_date=tran_date, spell_name="A",
                in_from=SPELL_A_IN_FROM, in_to=SPELL_A_IN_TO,
                spell_start=SPELL_A_START, spell_end=SPELL_A_END,
                is_off_day=is_off_day,
            )
            b_reg, b_ot = _process_one_spell(
                db, tran_date=tran_date, spell_name="B",
                in_from=SPELL_B_IN_FROM, in_to=SPELL_B_IN_TO,
                spell_start=SPELL_B_START, spell_end=SPELL_B_END,
                is_off_day=is_off_day,
            )
            db.commit()
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            raise HTTPException(
                status_code=500,
                detail=f"Spell processing failed: {e}",
            )

        total = a_reg + a_ot + b_reg + b_ot
        message = (
            f"Processed {total} row(s) for {tran_date} "
            f"(A: {a_reg} regular + {a_ot} OT, "
            f"B: {b_reg} regular + {b_ot} OT, "
            f"day_off: {is_off_day}, "
            f"unmatched_emp_codes: {len(unique_codes)})."
        )

        # If any emp_codes were unmatched, stream the xlsx report alongside
        # success — stats travel in headers so the UI can still show them.
        if unmatched:
            xlsx_bytes = _build_unmatched_xlsx(unmatched)
            return Response(
                content=xlsx_bytes,
                status_code=200,
                media_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
                headers={
                    "Content-Disposition": (
                        f'attachment; filename="unmatched_emp_codes_{tran_date}.xlsx"'
                    ),
                    "Access-Control-Expose-Headers": (
                        "Content-Disposition, X-Unmatched-Count, X-Unmatched-Rows, "
                        "X-Processed, X-Process-Message"
                    ),
                    "X-Unmatched-Count": str(len(unique_codes)),
                    "X-Unmatched-Rows": str(len(unmatched)),
                    "X-Processed": str(total),
                    "X-Process-Message": message,
                },
            )

        return {
            "message": message,
            "processed": total,
            "tran_date": tran_date,
            "is_off_day": is_off_day,
            "spell_a_regular": a_reg,
            "spell_a_ot": a_ot,
            "spell_b_regular": b_reg,
            "spell_b_ot": b_ot,
            "unmatched_emp_codes": 0,
        }
    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        # Log the full traceback so we can see it in uvicorn console.
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Final Process: copy daily_attendance_process_table -> daily_attendance.
# ---------------------------------------------------------------------------
# Spell rule (per user spec):
#   * Default: spell A if 06:00 <= ref_time < 14:00, spell B if 14:00 <= ref_time < 22:00.
#   * ref_time = check_out when ot_hours > 0, else check_in.
# Only rows with processed = 1 are copied; copied rows are marked processed = 2.

FINAL_FETCH_SQL = text(
    """
    SELECT daily_att_proc_id, eb_id, bio_id, dept_id, desig_id,
           attendance_date, spell_name, attendance_type, attendance_source,
           check_in, check_out, Time_duration, Working_hours, Ot_hours,
           spell_start_time, spell_end_time, spell_hours, device_name
      FROM daily_attendance_process_table
     WHERE attendance_date = :tran_date
       AND processed = 1
    """
)

FINAL_INSERT_SQL = text(
    """
    INSERT INTO daily_attendance
      (attendance_date, attendance_source, attendance_type, attendance_mark,
       eb_id, bio_id, branch_id, status_id,
       worked_department_id, worked_designation_id,
       entry_time, exit_time,
       working_hours, idle_hours, spell, spell_hours,
       is_active, update_date_time)
    VALUES
      (:attendance_date, :attendance_source, :attendance_type, 3,
       :eb_id, :bio_id, :branch_id, 1,
       :worked_department_id, :worked_designation_id,
       :entry_time, :exit_time,
       :working_hours, 0, :spell, :spell_hours,
       1, NOW())
    """
)

FINAL_MARK_PROCESSED_SQL = text(
    """
    UPDATE daily_attendance_process_table
       SET processed = 2
     WHERE attendance_date = :tran_date
       AND processed = 1
    """
)

FINAL_DELETE_EXISTING_SQL = text(
    """
    DELETE FROM daily_attendance
     WHERE bio_id IN :bio_ids
    """
)

# Fetch last daily_ebmc_attendance record for an eb_id, joined to
# daily_attendance so we can compare dept/desig.
FINAL_LAST_EBMC_SQL = text(
    """
    SELECT dea.mc_id,
           da.worked_department_id AS dept_id,
           da.worked_designation_id AS desig_id
    FROM daily_ebmc_attendance dea
    JOIN daily_attendance da ON da.daily_atten_id = dea.daily_atten_id
    WHERE dea.eb_id = :eb_id
      AND COALESCE(dea.is_active, 1) = 1
    ORDER BY dea.dtl_rec_id DESC
    LIMIT 1
    """
)

# Insert one mc entry into daily_ebmc_attendance.
FINAL_INSERT_EBMC_SQL = text(
    """
    INSERT INTO daily_ebmc_attendance
        (daily_atten_id, eb_id, mc_id, is_active)
    VALUES
        (:daily_atten_id, :eb_id, :mc_id, 1)
    """
)


def _resolve_spell_by_time(check_in, check_out, ot_hours) -> str | None:
    """Return 'A' / 'B' based on time-of-day rule.

    ref = check_out when ot_hours > 0 else check_in.
    A: 06:00 <= ref < 14:00 ; otherwise B.
    """
    try:
        ot_val = float(ot_hours or 0)
    except Exception:
        ot_val = 0.0
    ref = check_out if ot_val > 0 else check_in
    if ref is None:
        return None
    # Normalise to time-of-day seconds.
    if isinstance(ref, datetime):
        secs = ref.hour * 3600 + ref.minute * 60 + ref.second
    elif isinstance(ref, dt_time):
        secs = ref.hour * 3600 + ref.minute * 60 + ref.second
    elif isinstance(ref, timedelta):
        total = int(ref.total_seconds()) % 86400
        secs = total
    else:
        # Last-ditch: try to parse "HH:MM[:SS]".
        try:
            parts = str(ref).split(":")
            secs = int(parts[0]) * 3600 + int(parts[1]) * 60 + (int(parts[2]) if len(parts) > 2 else 0)
        except Exception:
            return None
    if 6 * 3600 <= secs < 14 * 3600:
        return "A"
    return "B"


@router.post("/bio_att_final_process")
async def bio_att_final_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Move processed=1 rows from daily_attendance_process_table into
    daily_attendance for a given date, recomputing spell by time-of-day.
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body = await request.json()
    except Exception:
        body = {}
    tran_date = (body or {}).get("tran_date")
    if not tran_date:
        raise HTTPException(status_code=400, detail="tran_date is required")
    try:
        datetime.strptime(tran_date, "%Y-%m-%d")
    except Exception:
        raise HTTPException(status_code=400, detail="tran_date must be YYYY-MM-DD")

    branch_id_raw = (body or {}).get("branch_id")
    if branch_id_raw in (None, ""):
        raise HTTPException(status_code=400, detail="branch_id is required")
    try:
        branch_id = int(branch_id_raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="branch_id must be an integer")

    try:
        _log_sql("FINAL_FETCH_SQL", FINAL_FETCH_SQL, {"tran_date": tran_date})
        rows = db.execute(FINAL_FETCH_SQL, {"tran_date": tran_date}).fetchall()
        print(f"[bio_att_final_process] fetched {len(rows)} row(s)", flush=True)

        # ── Delete existing daily_attendance rows for the same bio_ids ──────
        bio_ids = [r._mapping["bio_id"] for r in rows if r._mapping.get("bio_id") is not None]
        deleted_existing = 0
        if bio_ids:
            del_result = db.execute(FINAL_DELETE_EXISTING_SQL, {"bio_ids": tuple(bio_ids)})
            deleted_existing = del_result.rowcount
            print(
                f"[bio_att_final_process] deleted {deleted_existing} existing "
                f"daily_attendance row(s) for {len(bio_ids)} bio_id(s)",
                flush=True,
            )

        inserted = 0
        skipped = 0
        for r in rows:
            m = r._mapping
            spell = _resolve_spell_by_time(
                m.get("check_in"), m.get("check_out"), m.get("Ot_hours")
            )
            if spell is None:
                skipped += 1
                print(
                    f"[bio_att_final_process] skip eb_id={m.get('eb_id')} "
                    f"check_in={m.get('check_in')} check_out={m.get('check_out')} "
                    f"ot={m.get('Ot_hours')} {spell} (out of A/B window)",
                    flush=True,
                )
                continue

            base_params = {
                "attendance_date": m.get("attendance_date"),
                "attendance_source": m.get("attendance_source") or "BIO",
                "eb_id": m.get("eb_id"),
                "bio_id": m.get("bio_id"),
                "branch_id": branch_id,
                "worked_department_id": m.get("dept_id"),
                "worked_designation_id": m.get("desig_id"),
                "entry_time": m.get("check_in"),
                "exit_time": m.get("check_out"),
                "spell": spell,
                "spell_hours": m.get("spell_hours"),
            }

            try:
                wh = float(m.get("Working_hours") or 0)
            except Exception:
                wh = 0.0
            try:
                ot = float(m.get("Ot_hours") or 0)
            except Exception:
                ot = 0.0

            inserts: list[tuple[str, float]] = []
            if wh > 0 and ot > 0:
                inserts.append(("P", wh))
                inserts.append(("O", ot))
            elif wh > 0:
                inserts.append(("P", wh))
            elif ot > 0:
                inserts.append(("O", ot))
            else:
                skipped += 1
                print(
                    f"[bio_att_final_process] skip eb_id={m.get('eb_id')} "
                    f"working_hours=0 ot_hours=0",
                    flush=True,
                )
                continue

            for att_type, hours in inserts:
                params = {
                    **base_params,
                    "attendance_type": att_type,
                    "working_hours": hours,
                }
                _log_sql(
                    f"FINAL_INSERT_SQL [eb_id={m.get('eb_id')} spell={spell} type={att_type}]",
                    FINAL_INSERT_SQL, params,
                )
                ins_res = db.execute(FINAL_INSERT_SQL, params)
                inserted += 1

                # ── Insert daily_ebmc_attendance if last mc matches dept/desig ──
                new_daily_atten_id = ins_res.lastrowid
                eb_id_val = m.get("eb_id")
                curr_dept  = m.get("dept_id")
                curr_desig = m.get("desig_id")
                if new_daily_atten_id and eb_id_val and curr_dept and curr_desig:
                    last_ebmc = db.execute(
                        FINAL_LAST_EBMC_SQL, {"eb_id": eb_id_val}
                    ).fetchone()
                    if (
                        last_ebmc is not None
                        and last_ebmc.dept_id  is not None
                        and last_ebmc.desig_id is not None
                        and int(last_ebmc.dept_id)  == int(curr_dept)
                        and int(last_ebmc.desig_id) == int(curr_desig)
                    ):
                        db.execute(FINAL_INSERT_EBMC_SQL, {
                            "daily_atten_id": new_daily_atten_id,
                            "eb_id": eb_id_val,
                            "mc_id": last_ebmc.mc_id,
                        })
                        print(
                            f"[bio_att_final_process] ebmc inserted "
                            f"eb_id={eb_id_val} mc_id={last_ebmc.mc_id} "
                            f"daily_atten_id={new_daily_atten_id}",
                            flush=True,
                        )

        _log_sql("FINAL_MARK_PROCESSED_SQL", FINAL_MARK_PROCESSED_SQL, {"tran_date": tran_date})
        db.execute(FINAL_MARK_PROCESSED_SQL, {"tran_date": tran_date})
        db.commit()

        return {
            "message": (
                f"Final processed {inserted} row(s) for {tran_date} "
                f"(skipped {skipped} out-of-window, deleted {deleted_existing} existing)."
            ),
            "tran_date": tran_date,
            "inserted": inserted,
            "skipped": skipped,
            "deleted_existing": deleted_existing,
        }
    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Etrack SQL Server -> bio_attendance_table transfer
# ---------------------------------------------------------------------------

ETRACK_INSERT_SQL = text(
    """
    INSERT INTO bio_attendance_table
        (bio_att_log_id, emp_code, emp_anme, bio_id, log_date,
         device_direction, device_id)
    VALUES (:bio_att_log_id, :emp_code, :emp_anme, :bio_id, :log_date,
            :device_direction, :device_id)
    """
)


@router.post("/bio_att_etrack")
async def bio_att_etrack(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Transfer of punches from the Etrack SQL Server to bio_attendance_table.

    Two modes:

    1. Single-day mode — when `tran_date` is supplied:
           Fetches every row whose CONVERT(DATE, dl.LogDate) = tran_date from
           that day's monthly DeviceLogs_<m>_<y> table. No lower bound from
           prior records is applied — the entire selected day is pulled.

    2. Auto-incremental mode — when `tran_date` is omitted:
           The transfer range starts from MAX(log_date) in bio_attendance_table
           and spans every monthly DeviceLogs_<m>_<y> table from that month up
           to today's month, capped at dl.LogDate < (today + 1).
               - Start month  : dl.LogDate > last_log_date
               - Later months : all rows (still capped at today + 1)
           If bio_attendance_table is empty, only today's month is pulled.

    Body / query param (optional):
        company_id : Etrack CompanyId to filter on (default = 2)
        tran_date  : YYYY-MM-DD — switches to single-day mode

    De-dup key on the MySQL side: bio_att_log_id (= SQL Server DeviceLogId).
    """
    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            body = {}
        qp = request.query_params

        company_id_raw = body.get("company_id") or qp.get("company_id") or "2"
        try:
            company_id = int(company_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Invalid company_id")

        # Optional override for the upper bound of the transfer range. When
        # omitted or blank, falls back to today.
        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        tran_date_override: date | None = None
        if tran_date_raw:
            try:
                tran_date_override = datetime.strptime(
                    str(tran_date_raw), "%Y-%m-%d",
                ).date()
            except ValueError:
                raise HTTPException(
                    status_code=400,
                    detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
                )

        # Lazy import so the rest of the module still works without pyodbc.
        try:
            from src.hrms.etrack_conn import (
                device_logs_table_name,
                get_etrack_connection,
            )
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Etrack connector not available: {e}",
            )

        # ------------------------------------------------------------------
        # Determine the auto-incremental start point.
        # ------------------------------------------------------------------
        today = tran_date_override or date.today()

        last_log_row = db.execute(
            text(
                "SELECT MAX(log_date) AS max_log_date "
                "FROM bio_attendance_table"
            )
        ).mappings().first()
        last_log_date_dt = last_log_row["max_log_date"] if last_log_row else None

        if last_log_date_dt is None:
            # No prior data: start from selected day's month, take everything
            # in it (still capped at end_exclusive below).
            start_month_date = today.replace(day=1)
            last_log_dt: datetime | None = None
            from_log_date_iso: str | None = None
        else:
            if isinstance(last_log_date_dt, datetime):
                last_log_dt = last_log_date_dt
                last_log_d = last_log_date_dt.date()
            else:
                last_log_dt = datetime.combine(last_log_date_dt, dt_time.min)
                last_log_d = last_log_date_dt
            start_month_date = last_log_d.replace(day=1)
            from_log_date_iso = (
                last_log_date_dt.isoformat()
                if hasattr(last_log_date_dt, "isoformat")
                else str(last_log_date_dt)
            )

        # When a specific tran_date is supplied, fetch only that single day
        # from its monthly table, filtered by dl.DeviceLogId > last_id_for_day
        # so we incrementally pick up new punches for that date.
        # Otherwise, build the list of months from start_month_date through
        # today and let the per-month filters (last_log_dt + end_exclusive)
        # define the range.
        single_day_mode = tran_date_override is not None
        last_log_id_for_day = 0
        if single_day_mode:
            id_row = db.execute(
                text(
                    "SELECT MAX(bio_att_log_id) AS max_id "
                    "FROM bio_attendance_table "
                    "WHERE DATE(log_date) = :d"
                ),
                {"d": today},
            ).mappings().first()
            last_log_id_for_day = (
                int(id_row["max_id"])
                if id_row and id_row["max_id"] is not None
                else 0
            )

        months: list[date] = []
        if single_day_mode:
            months.append(today.replace(day=1))
        else:
            cursor_month = start_month_date
            end_month = today.replace(day=1)
            while cursor_month <= end_month:
                months.append(cursor_month)
                if cursor_month.month == 12:
                    cursor_month = date(cursor_month.year + 1, 1, 1)
                else:
                    cursor_month = date(
                        cursor_month.year, cursor_month.month + 1, 1
                    )

        # ------------------------------------------------------------------
        # Fetch from each monthly DeviceLogs table.
        # ------------------------------------------------------------------
        try:
            sconn = get_etrack_connection()
        except Exception as e:
            raise HTTPException(
                status_code=502,
                detail=f"Cannot connect to Etrack SQL Server: {e}",
            )

        total_fetched = 0
        total_inserted = 0
        per_table: list[dict] = []

        # Upper-bound LogDate (exclusive) — caps the transfer at the end of
        # the selected day (tran_date_override) or today.
        end_exclusive = today + timedelta(days=1)

        # ------------------------------------------------------------------
        # CSV dump of every row fetched from SQL Server (raw, before any
        # transformation). One file per request under exports/etrack/.
        # ------------------------------------------------------------------
        co_id_for_path = (qp.get("co_id") or "unknown").replace("/", "_")
        csv_dir = os.path.join("exports", "etrack")
        os.makedirs(csv_dir, exist_ok=True)
        csv_filename = (
            f"{co_id_for_path}_{today.isoformat()}_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        csv_path = os.path.abspath(os.path.join(csv_dir, csv_filename))
        csv_file = open(csv_path, "w", newline="", encoding="utf-8")
        csv_writer = csv.writer(csv_file)
        csv_writer.writerow([
            "DeviceLogId", "DeviceId", "UserId", "LogDate", "Direction",
            "EmployeeId", "EmployeeCode", "EmployeeName", "CompanyId",
            "source_table",
        ])

        try:
            cur = sconn.cursor()
            for idx, m_date in enumerate(months):
                table_name = device_logs_table_name(m_date)
                is_start_month = (idx == 0)

                if single_day_mode:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE em.CompanyId = ? "
                        f"  AND dl.DeviceLogId > ? "
                        f"  AND CONVERT(DATE, dl.LogDate) = ?"
                    )
                    params_args: tuple = (company_id, last_log_id_for_day, today)
                elif is_start_month and last_log_dt is not None:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE dl.LogDate > ? "
                        f"  AND em.CompanyId = ? "
                        f"  AND dl.LogDate < ?"
                    )
                    params_args = (last_log_dt, company_id, end_exclusive)
                else:
                    sql = (
                        f"SELECT dl.DeviceLogId, dl.DeviceId, dl.UserId, dl.LogDate, "
                        f"       dl.Direction, em.EmployeeId, em.EmployeeCode, "
                        f"       em.EmployeeName, em.CompanyId "
                        f"FROM dbo.{table_name} dl "
                        f"LEFT JOIN dbo.Employees em "
                        f"  ON em.EmployeeCodeInDevice = dl.UserId "
                        f"WHERE em.CompanyId = ? "
                        f"  AND dl.LogDate < ?"
                    )
                    params_args = (company_id, end_exclusive)
                try:
                    cur.execute(sql, *params_args)
                    src_rows = cur.fetchall()
                except Exception as e:
                    raise HTTPException(
                        status_code=500,
                        detail=f"Etrack query failed on {table_name}: {e}",
                    )
                print(sql)
                print(f"[bio_att_etrack] Executed on {table_name} with params {params_args}", flush=True)
                fetched = len(src_rows)
                print(f"[bio_att_etrack]  == {table_name}: fetched {fetched} row(s)", flush=True)
                inserted = 0
                BATCH_SIZE = 500
                batch: list[dict] = []

                def _flush_batch() -> int:
                    if not batch:
                        return 0
                    try:
                        res = db.execute(ETRACK_INSERT_SQL, batch)
                        flushed = int(res.rowcount or 0)
                        if flushed < 0:
                            flushed = len(batch)
                    except Exception as e:
                        print(
                            f"[bio_att_etrack] batch insert failed ({len(batch)} rows) on "
                            f"{table_name}: {e}. Falling back to row-by-row.",
                            flush=True,
                        )
                        flushed = 0
                        for p in batch:
                            try:
                                r2 = db.execute(ETRACK_INSERT_SQL, p)
                                flushed += int(r2.rowcount or 0)
                            except Exception as e2:
                                print(
                                    f"[bio_att_etrack] insert failed for "
                                    f"DeviceLogId={p['bio_att_log_id']}: {e2}",
                                    flush=True,
                                )
                    batch.clear()
                    return flushed

                for r in src_rows:
                    # CSV: dump every fetched row as-is (raw values).
                    csv_writer.writerow([
                        r.DeviceLogId, r.DeviceId, r.UserId, r.LogDate,
                        r.Direction, r.EmployeeId, r.EmployeeCode,
                        r.EmployeeName, r.CompanyId, table_name,
                    ])

                    params = {
                        "bio_att_log_id": r.DeviceLogId,
                        "emp_code": r.EmployeeCode,
                        "emp_anme": r.EmployeeName,
                        "bio_id": r.UserId,
                        "log_date": r.LogDate,
                        "device_direction": r.Direction,
                        "device_id": r.DeviceId,
                    }
                    if params["bio_att_log_id"] is None:
                        continue
                    batch.append(params)
                    if len(batch) >= BATCH_SIZE:
                        inserted += _flush_batch()

                inserted += _flush_batch()
                print(f"[bio_att_etrack] {table_name}: inserted {inserted} row(s)", flush=True)
                total_fetched += fetched
                total_inserted += inserted
                if single_day_mode:
                    per_from = today.isoformat()
                elif is_start_month:
                    per_from = from_log_date_iso
                else:
                    per_from = None
                per_table.append({
                    "table": table_name,
                    "from_log_date": per_from,
                    "fetched": fetched,
                    "inserted": inserted,
                })
        finally:
            try:
                sconn.close()
            except Exception:
                pass
            try:
                csv_file.close()
            except Exception:
                pass

        db.commit()
        duplicates = max(total_fetched - total_inserted, 0)

        # Step 1: back-fill eb_id (match_type='E') and device_id (match_type='B')
        # from tbl_master_bio_link_mst — same approach as the Excel upload flow.
        link_counts: dict = {}
        try:
            link_counts = _backfill_links(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            link_counts = {"errors": ["backfill failed"]}

        # Step 2: fill dept_id / desig_id from daily_attendance last record,
        # fallback to hrms_ed_official_details, for any rows with eb_id set
        # but dept_id / desig_id still NULL after step 1.
        dept_desig: dict = {}
        try:
            dept_desig = _resolve_dept_desig(db)
            db.commit()
        except Exception:
            try:
                db.rollback()
            except Exception:
                pass
            dept_desig = {"errors": ["dept_desig resolve failed"]}

        # First entry of `per_table` corresponds to the start month — keep
        # `source_table` for backwards-compatibility with existing UI.
        source_table = per_table[0]["table"] if per_table else ""

        return {
            "status": "ok",
            "from_log_date": (
                today.isoformat() if single_day_mode else from_log_date_iso
            ),
            "tran_date": today.isoformat(),
            "company_id": company_id,
            "source_table": source_table,
            "tables": per_table,
            "fetched": total_fetched,
            "inserted": total_inserted,
            "duplicates": duplicates,
            "csv_path": csv_path,
            "linked": {
                "eb_id": link_counts.get("E", 0),
                "device_id": link_counts.get("B", 0),
                "errors": link_counts.get("errors", []),
            },
            "dept_desig": {
                "candidates": dept_desig.get("candidates", 0),
                "updated": dept_desig.get("updated", 0),
                "from_daily_attendance": dept_desig.get("from_daily_attendance", 0),
                "fallback_official": dept_desig.get("fallback_official", 0),
                "no_source": dept_desig.get("no_source", 0),
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# ---------------------------------------------------------------------------
# Etrack Process: resolve eb_id / dept_id / desig_id for bio_attendance rows
# where eb_id IS NULL, using:
#   1. tbl_master_bio_link_mst (match_type='E') to resolve emp_code -> eb_id
#   2. last record in daily_attendance for that eb_id
#        -> worked_department_id (dept_id), worked_designation_id (desig_id)
#   3. fallback: hrms_ed_official_details
#        -> sub_dept_id (dept_id), designation_id (desig_id)
# ---------------------------------------------------------------------------

_ETRACK_PROC_UNRESOLVED_SQL = text(
    """
    SELECT DISTINCT b.emp_code, CAST(m.master_id AS SIGNED) AS eb_id
    FROM bio_attendance_table b
    JOIN tbl_master_bio_link_mst m
        ON m.match_type = 'E' AND m.bio_data = b.emp_code
    WHERE b.eb_id IS NULL
      AND b.emp_code IS NOT NULL AND b.emp_code <> ''
    """
)

_ETRACK_PROC_LAST_DAILY_ATT_SQL = text(
    """
    SELECT da.eb_id,
           da.worked_department_id,
           da.worked_designation_id
    FROM daily_attendance da
    INNER JOIN (
        SELECT eb_id, MAX(attendance_date) AS last_date
        FROM daily_attendance
        WHERE eb_id IN :eb_ids
          AND worked_department_id IS NOT NULL
        GROUP BY eb_id
    ) lx ON lx.eb_id = da.eb_id AND lx.last_date = da.attendance_date
    WHERE da.eb_id IN :eb_ids
      AND da.worked_department_id IS NOT NULL
    """
)

_ETRACK_PROC_OFFICIAL_SQL = text(
    """
    SELECT eb_id, sub_dept_id AS dept_id, designation_id AS desig_id
    FROM hrms_ed_official_details
    WHERE eb_id IN :eb_ids
      AND active = 1
    """
)

_ETRACK_PROC_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table
       SET eb_id    = :eb_id,
           dept_id  = :dept_id,
           desig_id = :desig_id
     WHERE emp_code = :emp_code
       AND eb_id IS NULL
    """
)

# ---------------------------------------------------------------------------
# Resolve dept_id / desig_id for rows that already have eb_id set but are
# still missing dept_id or desig_id.
# Priority:
#   1. last record in daily_attendance (worked_department_id, worked_designation_id)
#   2. fallback: hrms_ed_official_details (sub_dept_id, designation_id)
# ---------------------------------------------------------------------------

_RESOLVE_DEPT_DESIG_TARGET_SQL = text(
    """
    SELECT DISTINCT b.emp_code, CAST(b.eb_id AS SIGNED) AS eb_id
    FROM bio_attendance_table b
    WHERE b.eb_id IS NOT NULL
      AND (b.dept_id IS NULL OR b.desig_id IS NULL)
    """
)

_RESOLVE_DEPT_DESIG_UPDATE_SQL = text(
    """
    UPDATE bio_attendance_table
       SET dept_id  = :dept_id,
           desig_id = :desig_id
     WHERE emp_code = :emp_code
       AND eb_id    = :eb_id
       AND (dept_id IS NULL OR desig_id IS NULL)
    """
)


def _resolve_dept_desig(db: Session) -> dict:
    """Fill dept_id / desig_id on bio_attendance rows that have eb_id set
    but still have dept_id or desig_id NULL.

    Priority:
      1. Last record in daily_attendance per eb_id
           -> worked_department_id (dept_id), worked_designation_id (desig_id)
      2. Fallback: hrms_ed_official_details
           -> sub_dept_id (dept_id), designation_id (desig_id)

    Returns a stats dict with keys:
        candidates, updated, from_daily_attendance, fallback_official, no_source
    """
    result: dict = {
        "candidates": 0,
        "updated": 0,
        "from_daily_attendance": 0,
        "fallback_official": 0,
        "no_source": 0,
    }

    target_rows = db.execute(_RESOLVE_DEPT_DESIG_TARGET_SQL).fetchall()
    if not target_rows:
        return result

    emp_eb_map: dict[str, int] = {
        str(r.emp_code): int(r.eb_id)
        for r in target_rows
        if r.eb_id is not None
    }
    if not emp_eb_map:
        return result

    result["candidates"] = len(emp_eb_map)
    unique_eb_ids = list(set(emp_eb_map.values()))

    # Step 1: last daily_attendance record per eb_id
    da_rows = db.execute(
        _ETRACK_PROC_LAST_DAILY_ATT_SQL,
        {"eb_ids": tuple(unique_eb_ids)},
    ).fetchall()
    da_map: dict[int, tuple] = {
        int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
        for r in da_rows
        if r.worked_department_id is not None
    }
    result["from_daily_attendance"] = len(da_map)

    # Step 2: fallback from hrms_ed_official_details
    missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
    official_map: dict[int, tuple] = {}
    if missing_eb_ids:
        off_rows = db.execute(
            _ETRACK_PROC_OFFICIAL_SQL,
            {"eb_ids": tuple(missing_eb_ids)},
        ).fetchall()
        official_map = {
            int(r.eb_id): (r.dept_id, r.desig_id)
            for r in off_rows
        }

    updated = 0
    fallback_official = 0
    no_source = 0

    for emp_code, eb_id in emp_eb_map.items():
        if eb_id in da_map:
            dept_id, desig_id = da_map[eb_id]
        elif eb_id in official_map:
            dept_id, desig_id = official_map[eb_id]
            fallback_official += 1
        else:
            no_source += 1
            continue

        res = db.execute(
            _RESOLVE_DEPT_DESIG_UPDATE_SQL,
            {
                "dept_id": dept_id,
                "desig_id": desig_id,
                "emp_code": emp_code,
                "eb_id": eb_id,
            },
        )
        updated += int(res.rowcount or 0)

    result["updated"] = updated
    result["fallback_official"] = fallback_official
    result["no_source"] = no_source
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Bprocess-specific processing. Used by /bio_att_bprocess.
# Variant of etrack with different shift rules — cross-midnight C/C1 spell.
#
# Rule 1 — first_h in [5, 9]:
#   eff = last_h - 6
#   if eff >= 7        -> rec1 {spell:A,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:A1, type:R, working:4}
#   O = eff - rec1.working
#   if O >= 7          -> rec2 {spell:B,  type:O, ot:8}
#   elif O in [3,7)    -> rec2 {spell:B1, type:O, ot:4}
#
# Rule 2 — first_h in (9, 13]:
#   if span >= 3       -> rec1 {spell:A2, type:O, ot:4}
#   if span > 4:
#     d1 = last_h - 14
#     if d1 >= 7       -> rec2 {spell:B,  type:R, working:8}
#     elif d1 in [3,7) -> rec2 {spell:B1, type:R, working:4}
#     OB = d1 - rec2.working
#     if OB >= 7       -> rec3 {spell:C,  type:O, ot:8}
#     elif OB in [3,7) -> rec3 {spell:C1, type:O, ot:4}
#
# Rule 3 — first_h > 21 AND last on next day in [00:00, 05:00]:
#   last_ext = last_h + 24      (last on tran_date+1)
#   eff = last_ext - 22
#   if eff >= 7        -> rec1 {spell:C,  type:R, working:8}
#   elif eff in [3,7)  -> rec1 {spell:C1, type:R, working:4}
#   OC = eff - rec1.working
#   if OC >= 7         -> rec2 {spell:A,  type:O, ot:8}
#   elif OC in [3,7)   -> rec2 {spell:A1, type:O, ot:4}
# ─────────────────────────────────────────────────────────────────────────────

_BPROCESS_SPELL_TIMES: dict[str, tuple[str, str]] = {
    "A":  ("06:00:00", "14:00:00"),
    "A1": ("06:00:00", "14:00:00"),
    "A2": ("09:00:00", "13:00:00"),
    "B":  ("14:00:00", "22:00:00"),
    "B1": ("14:00:00", "22:00:00"),
    "C":  ("22:00:00", "06:00:00"),
    "C1": ("22:00:00", "06:00:00"),
}

FETCH_BPROCESS_PUNCHES_SQL = text(
    """
    SELECT b.eb_id,
           b.bio_att_log_id,
           b.dept_id,
           b.desig_id,
           DATE(b.log_date)   AS punch_date,
           TIME(b.log_date)   AS punch_time,
           b.log_date         AS log_date,
           b.device_direction AS device_direction
    FROM bio_attendance_table b
    WHERE b.eb_id     IS NOT NULL
      AND b.dept_id   IS NOT NULL
      AND b.desig_id  IS NOT NULL
      AND b.device_id IS NOT NULL
      AND (
            DATE(b.log_date) = :tran_date
         OR (DATE(b.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
             AND TIME(b.log_date) <= '06:00:00')
      )
    ORDER BY b.eb_id, b.log_date
    """
)


def _bprocess_records_for_employee(
    first_h: float, last_h: float, span: float, last_h_extended: float,
) -> list[dict]:
    out: list[dict] = []

    # Rule 1: first entry [5, 9]
    if 5 <= first_h <= 9:
        eff = last_h - 6
        rec1_working = 0.0
        if eff >= 7:
            out.append({"spell": "A",  "att_type": "R", "working": 8.0, "ot": 0.0})
            rec1_working = 8.0
        elif eff >= 3:
            out.append({"spell": "A1", "att_type": "R", "working": 4.0, "ot": 0.0})
            rec1_working = 4.0
        if rec1_working > 0:
            O = eff - rec1_working
            if O >= 7:
                out.append({"spell": "B",  "att_type": "O", "working": 0.0, "ot": 8.0})
            elif O >= 3:
                out.append({"spell": "B1", "att_type": "O", "working": 0.0, "ot": 4.0})

    # Rule 2: first entry (9, 13]
    elif 9 < first_h <= 13:
        if span >= 3:
            out.append({"spell": "A2", "att_type": "O", "working": 0.0, "ot": 4.0})
        if span > 4:
            d1 = last_h - 14
            rec2_working = 0.0
            if d1 >= 7:
                out.append({"spell": "B",  "att_type": "R", "working": 8.0, "ot": 0.0})
                rec2_working = 8.0
            elif d1 >= 3:
                out.append({"spell": "B1", "att_type": "R", "working": 4.0, "ot": 0.0})
                rec2_working = 4.0
            if rec2_working > 0:
                OB = d1 - rec2_working
                if OB >= 7:
                    out.append({"spell": "C",  "att_type": "O", "working": 0.0, "ot": 8.0})
                elif OB >= 3:
                    out.append({"spell": "C1", "att_type": "O", "working": 0.0, "ot": 4.0})

    # Rule 3: first entry > 21 with cross-midnight last
    elif first_h > 21 and last_h_extended >= 24:
        eff = last_h_extended - 22
        rec1_working = 0.0
        if eff >= 7:
            out.append({"spell": "C",  "att_type": "R", "working": 8.0, "ot": 0.0})
            rec1_working = 8.0
        elif eff >= 3:
            out.append({"spell": "C1", "att_type": "R", "working": 4.0, "ot": 0.0})
            rec1_working = 4.0
        if rec1_working > 0:
            OC = eff - rec1_working
            if OC >= 7:
                out.append({"spell": "A",  "att_type": "O", "working": 0.0, "ot": 8.0})
            elif OC >= 3:
                out.append({"spell": "A1", "att_type": "O", "working": 0.0, "ot": 4.0})

    return out


# Step 1 — prior night-shift OUT carryover.
# If yesterday's first punch was after 21:00, that employee is a night-shift
# worker, and their OUT happens on tran_date in [00:00..06:00]. Mark the LAST
# such punch as 'out' so it doesn't get treated as today's first IN.
_MARK_PRIOR_NIGHT_OUT_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT t.eb_id, MAX(t.log_date) AS edge_log
        FROM bio_attendance_table t
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_log
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = DATE_SUB(:tran_date, INTERVAL 1 DAY)
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) y ON t.eb_id = y.eb_id
        WHERE DATE(t.log_date) = :tran_date
          AND TIME(t.log_date) <= '06:00:00'
          AND HOUR(y.first_log) > 21
        GROUP BY t.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 2 — first IN of tran_date: earliest punch NOT already marked 'out'.
_MARK_FIRST_IN_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT eb_id, MIN(log_date) AS edge_log
        FROM bio_attendance_table
        WHERE eb_id IS NOT NULL
          AND DATE(log_date) = :tran_date
          AND (device_direction IS NULL OR device_direction <> 'out')
        GROUP BY eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'in'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 3 — day-shift OUT: last punch on tran_date when the first IN's hour
# is <= 21. The first-IN row itself is excluded so a single-punch day stays
# as 'in' rather than being flipped to 'out'.
_MARK_LAST_OUT_DAY_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT p.eb_id, MAX(p.log_date) AS edge_log
        FROM bio_attendance_table p
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_in
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = :tran_date
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) f ON p.eb_id = f.eb_id
        WHERE DATE(p.log_date) = :tran_date
          AND HOUR(f.first_in) <= 21
          AND p.log_date <> f.first_in
        GROUP BY p.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = :tran_date
    """
)

# Step 4 — night-shift OUT: for employees whose first IN on tran_date is
# after 21:00, the OUT is the LAST punch on tran_date+1 within [00:00..06:00].
_MARK_LAST_OUT_NIGHT_SQL = text(
    """
    UPDATE bio_attendance_table b
    JOIN (
        SELECT n.eb_id, MAX(n.log_date) AS edge_log
        FROM bio_attendance_table n
        JOIN (
            SELECT eb_id, MIN(log_date) AS first_in
            FROM bio_attendance_table
            WHERE eb_id IS NOT NULL
              AND DATE(log_date) = :tran_date
              AND (device_direction IS NULL OR device_direction <> 'out')
            GROUP BY eb_id
        ) f ON n.eb_id = f.eb_id
        WHERE DATE(n.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
          AND TIME(n.log_date) <= '06:00:00'
          AND HOUR(f.first_in) > 21
        GROUP BY n.eb_id
    ) m ON b.eb_id = m.eb_id AND b.log_date = m.edge_log
    SET b.device_direction = 'out'
    WHERE DATE(b.log_date) = DATE_ADD(:tran_date, INTERVAL 1 DAY)
    """
)


def _process_bprocess_day(
    db: Session,
    *,
    tran_date: str,
    is_off_day: bool,
) -> int:
    """Insert daily_attendance_process_table rows per the Bprocess rules.
    Punches from tran_date+1 [00:00..06:00] feed the night-shift Rule 3."""
    # Marking pipeline (order matters):
    #   1. Carry over yesterday's night-shift OUT into today's [00:00..06:00].
    #   2. Mark today's first IN as the earliest punch NOT already 'out'.
    #   3. Mark today's day-shift OUT (last punch when first IN <= 21).
    #   4. Mark night-shift OUT on tran_date+1 (when first IN > 21).
    db.execute(_MARK_PRIOR_NIGHT_OUT_SQL, {"tran_date": tran_date})
    db.execute(_MARK_FIRST_IN_SQL, {"tran_date": tran_date})
    db.execute(_MARK_LAST_OUT_DAY_SQL, {"tran_date": tran_date})
    db.execute(_MARK_LAST_OUT_NIGHT_SQL, {"tran_date": tran_date})

    rows = db.execute(
        FETCH_BPROCESS_PUNCHES_SQL, {"tran_date": tran_date},
    ).fetchall()

    by_emp: dict[int, list[tuple]] = {}
    for r in rows:
        m = r._mapping
        by_emp.setdefault(m["eb_id"], []).append((
            str(m["punch_date"]), m["punch_time"], m["bio_att_log_id"],
            m["dept_id"], m["desig_id"], m["log_date"],
            m["device_direction"],
        ))

    inserted = 0
    for eb_id, punches in by_emp.items():
        day0 = [p for p in punches if p[0] == tran_date]
        day1 = [p for p in punches if p[0] != tran_date]
        if not day0:
            continue

        # Skip leading day0 rows already marked 'out' — those are the prior
        # night shift's exit punches that landed in today's date.
        first_idx = 0
        while (
            first_idx < len(day0)
            and (day0[first_idx][6] or "").lower() == "out"
        ):
            first_idx += 1
        if first_idx >= len(day0):
            continue
        day0 = day0[first_idx:]

        first_pdate, first_time, first_bio, first_dept, first_desig, first_log_date, _ = day0[0]
        first_sec = _to_seconds(first_time)
        first_h   = first_sec / 3600.0

        # Pick last punch + extended hour based on which rule will fire.
        if first_h > 21 and day1:
            last_pdate, last_time, _, _, _, last_log_date, _ = day1[-1]
            last_sec = _to_seconds(last_time)
            last_h   = last_sec / 3600.0
            last_h_extended = last_h + 24.0
            total_secs = (last_sec + 24 * 3600) - first_sec
        else:
            last_pdate, last_time, _, _, _, last_log_date, _ = day0[-1]
            last_sec = _to_seconds(last_time)
            last_h   = last_sec / 3600.0
            last_h_extended = last_h
            total_secs = max(0, last_sec - first_sec)

        if total_secs < MIN_PAIR_SECONDS:
            continue

        span = round(last_h_extended - first_h, 4)

        recs = _bprocess_records_for_employee(first_h, last_h, span, last_h_extended)
        if not recs:
            continue

        for rec in recs:
            att_type      = "O" if is_off_day else rec["att_type"]
            working_hours = rec["working"]
            ot_hours      = rec["ot"]
            time_duration = round(working_hours + ot_hours, 2)
            spell_start, spell_end = _BPROCESS_SPELL_TIMES.get(
                rec["spell"], ("00:00:00", "00:00:00"),
            )
            db.execute(
                INSERT_SPELL_ROW_SQL,
                {
                    "eb_id": int(eb_id),
                    "bio_id": int(first_bio) if first_bio is not None else None,
                    "dept_id": int(first_dept) if first_dept is not None else None,
                    "desig_id": int(first_desig) if first_desig is not None else None,
                    "tran_date": tran_date,
                    "spell_name": rec["spell"],
                    "attendance_type": att_type,
                    "check_in": first_log_date,
                    "check_out": last_log_date,
                    "time_duration": time_duration,
                    "working_hours": working_hours,
                    "ot_hours": ot_hours,
                    "spell_start": spell_start,
                    "spell_end": spell_end,
                    "spell_hours": SPELL_HOURS,
                },
            )
            inserted += 1

    print(f"[bio_att_bprocess]   -> rows inserted = {inserted}", flush=True)
    return inserted


@router.post("/bio_att_bprocess")
async def bio_att_bprocess(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Same orchestration as /bio_att_etrack_process — resolve eb_id /
    dept_id / desig_id on bio_attendance_table, then build daily rows — but
    using the Bprocess rules (Rules #1/#2/#3 above, including cross-midnight
    C/C1 spell)."""
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        if not tran_date_raw:
            raise HTTPException(status_code=400, detail="tran_date is required")
        try:
            datetime.strptime(tran_date_raw, "%Y-%m-%d")
            tran_date: str = tran_date_raw
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
            )

        branch_id_raw = body.get("branch_id") or qp.get("branch_id")
        if branch_id_raw in (None, ""):
            raise HTTPException(status_code=400, detail="branch_id is required")
        try:
            branch_id = int(branch_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="branch_id must be an integer")

        unresolved_rows = db.execute(_ETRACK_PROC_UNRESOLVED_SQL).fetchall()

        resolve_result = {
            "resolved": 0,
            "updated": 0,
            "from_daily_attendance": 0,
            "fallback_official": 0,
            "no_source": 0,
        }

        if unresolved_rows:
            emp_eb_map: dict[str, int] = {
                str(r.emp_code): int(r.eb_id)
                for r in unresolved_rows
                if r.eb_id is not None
            }

            if emp_eb_map:
                unique_eb_ids = list(set(emp_eb_map.values()))

                da_rows = db.execute(
                    _ETRACK_PROC_LAST_DAILY_ATT_SQL,
                    {"eb_ids": tuple(unique_eb_ids)},
                ).fetchall()
                da_map: dict[int, tuple] = {
                    int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
                    for r in da_rows
                    if r.worked_department_id is not None
                }

                missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
                official_map: dict[int, tuple] = {}
                if missing_eb_ids:
                    off_rows = db.execute(
                        _ETRACK_PROC_OFFICIAL_SQL,
                        {"eb_ids": tuple(missing_eb_ids)},
                    ).fetchall()
                    official_map = {
                        int(r.eb_id): (r.dept_id, r.desig_id)
                        for r in off_rows
                    }

                updated = 0
                fallback_official = 0
                no_source = 0

                for emp_code, eb_id in emp_eb_map.items():
                    if eb_id in da_map:
                        dept_id, desig_id = da_map[eb_id]
                    elif eb_id in official_map:
                        dept_id, desig_id = official_map[eb_id]
                        fallback_official += 1
                    else:
                        no_source += 1
                        continue

                    res = db.execute(
                        _ETRACK_PROC_UPDATE_SQL,
                        {
                            "eb_id": eb_id,
                            "dept_id": dept_id,
                            "desig_id": desig_id,
                            "emp_code": emp_code,
                        },
                    )
                    updated += int(res.rowcount or 0)

                db.commit()

                resolve_result = {
                    "resolved": len(emp_eb_map),
                    "updated": updated,
                    "from_daily_attendance": len(da_map),
                    "fallback_official": fallback_official,
                    "no_source": no_source,
                }

        is_off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
        is_off_day = bool(is_off_row and int(is_off_row.cnt) > 0)

        db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
        db.commit()

        inserted = _process_bprocess_day(
            db, tran_date=tran_date, is_off_day=is_off_day,
        )
        db.commit()

        return {
            "status": "ok",
            "tran_date": tran_date,
            "branch_id": branch_id,
            "is_off_day": is_off_day,
            "resolve": resolve_result,
            "process": {
                "total_inserted": inserted,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


@router.post("/bio_att_etrack_process")
async def bio_att_etrack_process(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Resolve eb_id / dept_id / desig_id for bio_attendance rows where eb_id IS NULL,
    then process them into daily_attendance_process_table for the given date.

    Steps per unresolved emp_code:
      1. Look up eb_id from tbl_master_bio_link_mst (match_type='E').
      2. Check last record in daily_attendance for that eb_id:
           worked_department_id -> dept_id, worked_designation_id -> desig_id.
      3. If no daily_attendance record found, fall back to hrms_ed_official_details:
           sub_dept_id -> dept_id, designation_id -> desig_id.
      4. UPDATE bio_attendance_table for that emp_code (eb_id IS NULL rows only).
      5. Delete existing daily_attendance_process_table rows for tran_date,
         then run Spell A + Spell B processing into daily_attendance_process_table.

    Required body params:
        tran_date  : YYYY-MM-DD
        branch_id  : int
    """
    co_id = request.query_params.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    try:
        body: dict = {}
        try:
            body = await request.json()
        except Exception:
            pass
        qp = request.query_params

        # ── Validate required params ─────────────────────────────────────────
        tran_date_raw = body.get("tran_date") or qp.get("tran_date")
        if not tran_date_raw:
            raise HTTPException(status_code=400, detail="tran_date is required")
        try:
            datetime.strptime(tran_date_raw, "%Y-%m-%d")
            tran_date: str = tran_date_raw
        except ValueError:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid tran_date {tran_date_raw!r}, expected YYYY-MM-DD",
            )

        branch_id_raw = body.get("branch_id") or qp.get("branch_id")
        if branch_id_raw in (None, ""):
            raise HTTPException(status_code=400, detail="branch_id is required")
        try:
            branch_id = int(branch_id_raw)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="branch_id must be an integer")

        # ── Step 1: distinct emp_code -> eb_id from link master ─────────────
        unresolved_rows = db.execute(_ETRACK_PROC_UNRESOLVED_SQL).fetchall()

        resolve_result = {
            "resolved": 0,
            "updated": 0,
            "from_daily_attendance": 0,
            "fallback_official": 0,
            "no_source": 0,
        }

        if unresolved_rows:
            emp_eb_map: dict[str, int] = {
                str(r.emp_code): int(r.eb_id)
                for r in unresolved_rows
                if r.eb_id is not None
            }

            if emp_eb_map:
                unique_eb_ids = list(set(emp_eb_map.values()))

                # ── Step 2: last daily_attendance per eb_id ──────────────────
                da_rows = db.execute(
                    _ETRACK_PROC_LAST_DAILY_ATT_SQL,
                    {"eb_ids": tuple(unique_eb_ids)},
                ).fetchall()
                da_map: dict[int, tuple] = {
                    int(r.eb_id): (r.worked_department_id, r.worked_designation_id)
                    for r in da_rows
                    if r.worked_department_id is not None
                }

                # ── Step 3: fallback from hrms_ed_official_details ───────────
                missing_eb_ids = [eid for eid in unique_eb_ids if eid not in da_map]
                official_map: dict[int, tuple] = {}
                if missing_eb_ids:
                    off_rows = db.execute(
                        _ETRACK_PROC_OFFICIAL_SQL,
                        {"eb_ids": tuple(missing_eb_ids)},
                    ).fetchall()
                    official_map = {
                        int(r.eb_id): (r.dept_id, r.desig_id)
                        for r in off_rows
                    }

                # ── Step 4: UPDATE bio_attendance_table ──────────────────────
                updated = 0
                fallback_official = 0
                no_source = 0

                for emp_code, eb_id in emp_eb_map.items():
                    if eb_id in da_map:
                        dept_id, desig_id = da_map[eb_id]
                    elif eb_id in official_map:
                        dept_id, desig_id = official_map[eb_id]
                        fallback_official += 1
                    else:
                        no_source += 1
                        continue

                    res = db.execute(
                        _ETRACK_PROC_UPDATE_SQL,
                        {
                            "eb_id": eb_id,
                            "dept_id": dept_id,
                            "desig_id": desig_id,
                            "emp_code": emp_code,
                        },
                    )
                    updated += int(res.rowcount or 0)

                db.commit()

                resolve_result = {
                    "resolved": len(emp_eb_map),
                    "updated": updated,
                    "from_daily_attendance": len(da_map),
                    "fallback_official": fallback_official,
                    "no_source": no_source,
                }

        # ── Step 5: process into daily_attendance_process_table ─────────────
        is_off_row = db.execute(IS_OFF_DAY_SQL, {"tran_date": tran_date}).fetchone()
        is_off_day = bool(is_off_row and int(is_off_row.cnt) > 0)

        # Delete existing rows for the date (re-run safe).
        db.execute(DELETE_DAY_ROWS_SQL, {"tran_date": tran_date})
        db.commit()

        inserted = _process_etrack_day(
            db, tran_date=tran_date, is_off_day=is_off_day,
        )
        db.commit()

        return {
            "status": "ok",
            "tran_date": tran_date,
            "branch_id": branch_id,
            "is_off_day": is_off_day,
            "resolve": resolve_result,
            "process": {
                "total_inserted": inserted,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        try:
            db.rollback()
        except Exception:
            pass
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")


# =============================================================================
# Wages Register (attwgs) — pivoted shift + wages report
# =============================================================================
#
# Per (employee, attendance_date):
#   shift_letter  = 'A' if check_in hour in [5..12], 'B' if [13..20]
#   shift_code    = shift_letter             when working_hours == 8
#                   shift_letter + '1'       otherwise
#   rate          = latest employee_rate_table.rate where rate_date <= attendance_date
#   wages         = (rate / 8) * (working_hours + ot_hours)

_WAGES_REGISTER_SQL = text(
    """
                    SELECT
        d.eb_id                                                          AS eb_id,
        o.emp_code                                                       AS emp_code,
        TRIM(CONCAT_WS(' ', p.first_name,
                            IFNULL(p.middle_name, ''),
                            IFNULL(p.last_name,  ''))) AS emp_name,
        sdm.sub_dept_desc                                                    AS department,
        dm.desig                                                     AS designation,
        d.attendance_date                                                AS attendance_date,
        d.spell_name                                                     AS spell_name,
        d.attendance_type                                                AS attendance_type,
        d.check_in                                                       AS check_in,
        d.Working_hours                                                  AS working_hours,
        d.Ot_hours                                                       AS ot_hours,
        (
            SELECT er.rate
            FROM employee_rate_table er
            WHERE er.eb_id = d.eb_id
              AND er.rate_date <= d.attendance_date
            ORDER BY er.rate_date DESC
            LIMIT 1
        )                                                                AS rate
    FROM daily_attendance_process_table d
    LEFT JOIN hrms_ed_official_details o ON o.eb_id = d.eb_id
    LEFT JOIN hrms_ed_personal_details p ON p.eb_id = d.eb_id
	left join sub_dept_mst sdm on o.sub_dept_id =sdm.sub_dept_id 
    left join designation_mst dm on dm.designation_id =o.designation_id 
	WHERE d.attendance_date BETWEEN :from_date AND :to_date
      AND (:branch_id = 0 OR o.branch_id = :branch_id)
      AND (:emp_code = '' OR o.emp_code = :emp_code)
    ORDER BY CAST(o.emp_code AS UNSIGNED), o.emp_code, d.attendance_date
    """
)
 
def _shift_letter_from_check_in(check_in) -> str:
    """A if check_in hour in [5..12], B if in [13..20], else ''."""
    if check_in is None:
        return ""
    if isinstance(check_in, datetime):
        hour = check_in.hour
    elif isinstance(check_in, dt_time):
        hour = check_in.hour
    elif isinstance(check_in, timedelta):
        hour = int(check_in.total_seconds() // 3600) % 24
    else:
        try:
            hour = int(str(check_in).split(":")[0])
        except Exception:
            return ""
    if 5 <= hour <= 12:
        return "A"
    if 13 <= hour <= 20:
        return "B"
    return ""


@router.get("/wages_register")
async def wages_register(
    request: Request,
    db: Session = Depends(get_tenant_db),
    token_data: dict = Depends(get_current_user_with_refresh),
):
    """Pivoted wages register: shift code + per-day wages per employee."""
    qp = request.query_params
    co_id = qp.get("co_id")
    if not co_id:
        raise HTTPException(status_code=400, detail="co_id is required")

    from_raw = qp.get("from_date")
    to_raw = qp.get("to_date")
    if not from_raw or not to_raw:
        raise HTTPException(status_code=400, detail="from_date and to_date are required")
    try:
        from_date = datetime.strptime(from_raw, "%Y-%m-%d").date()
        to_date = datetime.strptime(to_raw, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="dates must be YYYY-MM-DD")
    if from_date > to_date:
        raise HTTPException(status_code=400, detail="from_date must be <= to_date")

    branch_raw = qp.get("branch_id") or "0"
    try:
        branch_id = int(branch_raw)
    except ValueError:
        raise HTTPException(status_code=400, detail="branch_id must be an integer")

    emp_code = (qp.get("emp_code") or "").strip()

    try:
        rows = db.execute(
            _WAGES_REGISTER_SQL,
            {
                "from_date": from_date,
                "to_date": to_date,
                "branch_id": branch_id,
                "emp_code": emp_code,
            },
        ).fetchall()

        columns: list[dict] = []
        d = from_date
        while d <= to_date:
            columns.append({
                "key": d.isoformat(),
                "label": d.strftime("%a, %b %d"),
            })
            d += timedelta(days=1)

        emp_meta:      dict[int, dict] = {}
        emp_shifts:    dict[int, dict[str, str]] = {}
        emp_ot_shifts: dict[int, dict[str, str]] = {}
        emp_wages:     dict[int, dict[str, float]] = {}
        emp_whrs_by:   dict[int, dict[str, float]] = {}
        emp_ot_by:     dict[int, dict[str, float]] = {}
        emp_whrs:      dict[int, float] = {}
        emp_ot:        dict[int, float] = {}
        emp_total:     dict[int, float] = {}
        emp_rate:      dict[int, float] = {}
        emp_p_days:    dict[int, set[str]] = {}
        emp_ot_days:   dict[int, set[str]] = {}

        for r in rows:
            m = r._mapping
            eb_id = m["eb_id"]
            if eb_id is None:
                continue
            if eb_id not in emp_meta:
                emp_meta[eb_id] = {
                    "eb_id": int(eb_id),
                    "emp_code": m["emp_code"] or "",
                    "emp_name": m["emp_name"] or "",
                    "department": m["department"] or "",
                    "designation": m["designation"] or "",
                }
                emp_shifts[eb_id]    = {}
                emp_ot_shifts[eb_id] = {}
                emp_wages[eb_id]     = {}
                emp_whrs_by[eb_id]   = {}
                emp_ot_by[eb_id]     = {}
                emp_whrs[eb_id]      = 0.0
                emp_ot[eb_id]        = 0.0
                emp_total[eb_id]     = 0.0
                emp_rate[eb_id]      = float(m["rate"] or 0)
                emp_p_days[eb_id]    = set()
                emp_ot_days[eb_id]   = set()

            att_date = m["attendance_date"]
            if att_date is None:
                continue
            if isinstance(att_date, datetime):
                att_date = att_date.date()
            key = att_date.isoformat()

            wh = float(m["working_hours"] or 0)
            ot = float(m["ot_hours"] or 0)
            rate = float(m["rate"] or 0)

            shift_from_check = _shift_letter_from_check_in(m["check_in"])
            shift_letter = shift_from_check or (m["spell_name"] or "")
            if not shift_letter:
                continue

            # Working hours -> Shift cell; OT hours -> OT cell. Each row of
            # daily_attendance_process_table can contribute to both. The "1" suffix
            # marks a partial-bucket value (hours != 8).
            if wh > 0:
                shift_code = shift_letter if wh == 8 else f"{shift_letter}1"
                existing = emp_shifts[eb_id].get(key, "")
                emp_shifts[eb_id][key] = (
                    f"{existing} {shift_code}".strip() if existing else shift_code
                )
            if ot > 0:
                ot_code = shift_letter if ot == 8 else f"{shift_letter}1"
                existing = emp_ot_shifts[eb_id].get(key, "")
                emp_ot_shifts[eb_id][key] = (
                    f"{existing} {ot_code}".strip() if existing else ot_code
                )

            wages = (rate / 8.0) * (wh + ot)
            emp_wages[eb_id][key] = round(
                emp_wages[eb_id].get(key, 0.0) + wages, 2
            )
            emp_whrs_by[eb_id][key] = round(
                emp_whrs_by[eb_id].get(key, 0.0) + wh, 2
            )
            emp_ot_by[eb_id][key] = round(
                emp_ot_by[eb_id].get(key, 0.0) + ot, 2
            )
            emp_whrs[eb_id]  += wh
            emp_ot[eb_id]    += ot
            emp_total[eb_id] += wages
            if wh > 0:
                emp_p_days[eb_id].add(key)
            if ot > 0:
                emp_ot_days[eb_id].add(key)
            if rate:
                emp_rate[eb_id] = rate

        out = []
        for eb_id, meta in emp_meta.items():
            out.append({
                **meta,
                "rate": round(emp_rate.get(eb_id, 0.0), 2),
                "shifts":         emp_shifts[eb_id],
                "ot_shifts":      emp_ot_shifts[eb_id],
                "wages":          emp_wages[eb_id],
                "working_hours":  emp_whrs_by[eb_id],
                "ot_hours":       emp_ot_by[eb_id],
                "total_working_hours": round(emp_whrs[eb_id], 2),
                "total_ot_hours":      round(emp_ot[eb_id],   2),
                "total_wages":         round(emp_total[eb_id], 2),
                "count_p_days":  len(emp_p_days[eb_id]),
                "count_ot_days": len(emp_ot_days[eb_id]),
            })
        # SQL already ORDER BY CAST(emp_code AS UNSIGNED), so dict-iteration order
        # of emp_meta preserves the desired numeric sort.
        return {"columns": columns, "data": out}
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")
